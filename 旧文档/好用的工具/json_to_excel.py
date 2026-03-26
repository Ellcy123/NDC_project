import json
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def json_to_excel(json_file, output_file):
    """将 JSON 文件转换为 Excel"""

    # 读取 JSON 文件
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 检查数据类型
    if not isinstance(data, list):
        print("错误: JSON 文件必须是数组格式")
        return False

    if len(data) == 0:
        print("错误: JSON 数据为空")
        return False

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # 获取所有字段名（从第一个对象）
    headers = list(data[0].keys())

    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据
    for row_idx, item in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = item.get(header, "")

            # 处理布尔值
            if isinstance(value, bool):
                value = str(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    # 对于中文字符，每个字符算2个长度
                    cell_length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, 80)  # 最大宽度限制为80
        ws.column_dimensions[column_letter].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存文件
    wb.save(output_file)
    print(f"Excel 文件已生成: {output_file}")
    print(f"共转换 {len(data)} 行数据, {len(headers)} 个字段")

    return True

if __name__ == "__main__":
    # 支持命令行参数
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        # 如果没有指定输出文件，自动生成（将 .json 替换为 .xlsx）
        if len(sys.argv) >= 3:
            output_file = sys.argv[2]
        else:
            output_file = os.path.splitext(input_file)[0] + '.xlsx'
    else:
        # 默认文件
        input_file = r"d:\NDC_project\Config\story_tbitemstaticdata.json"
        output_file = r"d:\NDC_project\Config\story_tbitemstaticdata.xlsx"

    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到输入文件: {input_file}")
        sys.exit(1)

    print(f"输入文件: {input_file}")
    print(f"输出文件: {output_file}")
    print("开始转换...")

    success = json_to_excel(input_file, output_file)

    if not success:
        sys.exit(1)
