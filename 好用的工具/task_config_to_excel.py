import json
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def task_config_to_excel(json_file, output_file):
    """将任务配置JSON文件转换为Excel"""

    # 读取 JSON 文件
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 检查数据结构
    if 'tasks' not in data:
        print("错误: JSON 文件必须包含 'tasks' 字段")
        return False

    tasks = data['tasks']

    if len(tasks) == 0:
        print("错误: tasks 数据为空")
        return False

    # 创建工作簿
    wb = openpyxl.Workbook()

    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ========== 工作表1: 任务主表 ==========
    ws_main = wb.create_sheet("任务主表")

    # 定义表头
    headers = [
        'taskId', 'taskType', 'sectionId', 'loopId', 'priority',
        'cnTitle', 'enTitle', 'cnDescription', 'enDescription',
        'showProgressBar', 'progressType', 'progressTotal', 'requiredEvidences',
        'showCompletionEffect', 'highlightWhen100Percent', 'highlight100PercentText_cn', 'highlight100PercentText_en',
        'maxDisplayCount', 'triggerCondition', 'completeCondition'
    ]

    header_cn = [
        '任务ID', '任务类型', '子章节ID', '循环ID', '优先级',
        '中文标题', '英文标题', '中文描述', '英文描述',
        '显示进度条', '进度类型', '进度总数', '需要的证据列表',
        '显示完成动效', '100%时高亮', '高亮提示_中文', '高亮提示_英文',
        '最大显示数量', '触发条件', '完成条件'
    ]

    # 写入中文表头（第1行）
    for col_idx, header_text in enumerate(header_cn, start=1):
        cell = ws_main.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 写入英文字段名（第2行）
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_main.cell(row=2, column=col_idx, value=header)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 写入数据
    for row_idx, task in enumerate(tasks, start=3):
        # 提取highlight100PercentText
        highlight_cn = ""
        highlight_en = ""
        if 'highlight100PercentText' in task and task['highlight100PercentText']:
            highlight_cn = task['highlight100PercentText'].get('cn', '')
            highlight_en = task['highlight100PercentText'].get('en', '')

        # 提取requiredEvidences
        required_evidences = ""
        if 'requiredEvidences' in task and task['requiredEvidences']:
            required_evidences = ", ".join(task['requiredEvidences'])

        row_data = [
            task.get('taskId', ''),
            task.get('taskType', ''),
            task.get('sectionId', ''),
            task.get('loopId', ''),
            task.get('priority', ''),
            task.get('cnTitle', ''),
            task.get('enTitle', ''),
            task.get('cnDescription', ''),
            task.get('enDescription', ''),
            task.get('showProgressBar', ''),
            task.get('progressType', ''),
            task.get('progressTotal', ''),
            required_evidences,
            task.get('showCompletionEffect', ''),
            task.get('highlightWhen100Percent', ''),
            highlight_cn,
            highlight_en,
            task.get('maxDisplayCount', ''),
            task.get('triggerCondition', ''),
            task.get('completeCondition', '')
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 根据任务类型设置不同的背景色
            task_type = task.get('taskType', '')
            if task_type == 'MainCase':
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif task_type == 'PhaseGoal':
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            elif task_type == 'CurrentGoal':
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            elif task_type == 'Doubt':
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            elif task_type == 'SideCase':
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # 自动调整列宽
    for column in ws_main.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    # 对于中文字符，每个字符算2个长度
                    cell_length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max(max_length + 2, 10), 60)  # 最小10，最大60
        ws_main.column_dimensions[column_letter].width = adjusted_width

    # 冻结前两行
    ws_main.freeze_panes = 'A3'

    # ========== 工作表2: 子任务表 ==========
    ws_subtask = wb.create_sheet("子任务表")

    subtask_headers = ['taskId', 'subtaskId', 'cnText', 'enText', 'completeCondition']
    subtask_headers_cn = ['父任务ID', '子任务ID', '中文描述', '英文描述', '完成条件']

    # 写入中文表头
    for col_idx, header_text in enumerate(subtask_headers_cn, start=1):
        cell = ws_subtask.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 写入英文字段名
    for col_idx, header in enumerate(subtask_headers, start=1):
        cell = ws_subtask.cell(row=2, column=col_idx, value=header)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 写入子任务数据
    subtask_row = 3
    for task in tasks:
        if 'subtasks' in task and task['subtasks']:
            for subtask in task['subtasks']:
                row_data = [
                    task.get('taskId', ''),
                    subtask.get('subtaskId', ''),
                    subtask.get('cnText', ''),
                    subtask.get('enText', ''),
                    subtask.get('completeCondition', '')
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_subtask.cell(row=subtask_row, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                subtask_row += 1

    # 自动调整列宽
    for column in ws_subtask.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    cell_length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max(max_length + 2, 10), 60)
        ws_subtask.column_dimensions[column_letter].width = adjusted_width

    ws_subtask.freeze_panes = 'A3'

    # ========== 工作表3: 任务类型配置 ==========
    ws_type = wb.create_sheet("任务类型配置")

    type_headers = ['taskType', 'maxDisplayCount', 'showProgressBar', 'showCompletionEffect', 'defaultPriority']
    type_headers_cn = ['任务类型', '最大显示数量', '显示进度条', '显示完成动效', '默认优先级']

    # 写入中文表头
    for col_idx, header_text in enumerate(type_headers_cn, start=1):
        cell = ws_type.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 写入英文字段名
    for col_idx, header in enumerate(type_headers, start=1):
        cell = ws_type.cell(row=2, column=col_idx, value=header)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 写入类型配置数据
    if 'taskTypeConfig' in data:
        type_row = 3
        for task_type, config in data['taskTypeConfig'].items():
            row_data = [
                task_type,
                config.get('maxDisplayCount', ''),
                config.get('showProgressBar', ''),
                config.get('showCompletionEffect', ''),
                config.get('defaultPriority', '')
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_type.cell(row=type_row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            type_row += 1

    # 自动调整列宽
    for col_idx in range(1, len(type_headers) + 1):
        ws_type.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    ws_type.freeze_panes = 'A3'

    # ========== 工作表4: 显示优先级 ==========
    ws_priority = wb.create_sheet("显示优先级")

    # 缩略展示优先级
    ws_priority.cell(row=1, column=1, value="缩略展示优先级").font = Font(bold=True, size=12)
    ws_priority.cell(row=2, column=1, value="顺序").font = Font(bold=True)
    ws_priority.cell(row=2, column=2, value="任务类型").font = Font(bold=True)

    if 'shortcutDisplayPriority' in data:
        for idx, task_type in enumerate(data['shortcutDisplayPriority'], start=3):
            ws_priority.cell(row=idx, column=1, value=idx-2)
            ws_priority.cell(row=idx, column=2, value=task_type)

    # 展开展示顺序
    ws_priority.cell(row=1, column=4, value="展开展示顺序").font = Font(bold=True, size=12)
    ws_priority.cell(row=2, column=4, value="顺序").font = Font(bold=True)
    ws_priority.cell(row=2, column=5, value="任务类型").font = Font(bold=True)

    if 'expandedDisplayOrder' in data:
        for idx, task_type in enumerate(data['expandedDisplayOrder'], start=3):
            ws_priority.cell(row=idx, column=4, value=idx-2)
            ws_priority.cell(row=idx, column=5, value=task_type)

    # 调整列宽
    ws_priority.column_dimensions['A'].width = 10
    ws_priority.column_dimensions['B'].width = 20
    ws_priority.column_dimensions['C'].width = 5
    ws_priority.column_dimensions['D'].width = 10
    ws_priority.column_dimensions['E'].width = 20

    # 保存文件
    wb.save(output_file)
    print(f"[OK] Excel 文件已生成: {output_file}")
    print(f"[INFO] 共转换 {len(tasks)} 个任务")
    print(f"[INFO] 工作表: 任务主表, 子任务表, 任务类型配置, 显示优先级")

    return True

if __name__ == "__main__":
    # 支持命令行参数
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        # 如果没有指定输出文件，自动生成
        if len(sys.argv) >= 3:
            output_file = sys.argv[2]
        else:
            output_file = os.path.splitext(input_file)[0] + '.xlsx'
    else:
        # 默认文件
        input_file = r"d:\NDC_project\Config\TaskConfig_Episode1_Loop1_Example.json"
        output_file = r"d:\NDC_project\Config\TaskConfig_Episode1_Loop1_Example.xlsx"

    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"[ERROR] 找不到输入文件: {input_file}")
        sys.exit(1)

    print(f"[INFO] 输入文件: {input_file}")
    print(f"[INFO] 输出文件: {output_file}")
    print("[INFO] 开始转换...")
    print()

    success = task_config_to_excel(input_file, output_file)

    if not success:
        sys.exit(1)
