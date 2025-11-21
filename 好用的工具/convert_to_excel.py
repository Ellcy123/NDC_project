import re
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def parse_markdown_table(md_file_path):
    """解析 Markdown 文件中的表格"""
    with open(md_file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 查找所有表格
    tables = []
    lines = content.split('\n')

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 检测表格开始（以 | 开头和结尾）
        if line.startswith('|') and line.endswith('|'):
            table_lines = []
            table_name = ""

            # 向上查找表格标题
            for j in range(i-1, max(0, i-5), -1):
                if lines[j].strip().startswith('###'):
                    table_name = lines[j].strip().replace('###', '').strip()
                    break
                elif lines[j].strip().startswith('##'):
                    table_name = lines[j].strip().replace('##', '').strip()
                    break

            # 收集表格所有行
            while i < len(lines) and lines[i].strip().startswith('|'):
                table_lines.append(lines[i].strip())
                i += 1

            if len(table_lines) >= 3:  # 至少需要表头、分隔符、数据行
                tables.append({
                    'name': table_name if table_name else f'表格{len(tables)+1}',
                    'lines': table_lines
                })
            continue

        i += 1

    return tables

def parse_table_lines(table_lines):
    """将表格行解析为数据"""
    data = []

    for i, line in enumerate(table_lines):
        # 跳过分隔符行（包含 :---: 或 --- 的行）
        if i == 1 and (':---' in line or '---' in line):
            continue

        # 分割单元格
        cells = [cell.strip() for cell in line.split('|')]
        # 移除首尾空元素
        cells = [c for c in cells if c]

        if cells:
            data.append(cells)

    return data

def create_excel(tables, output_file):
    """创建 Excel 文件"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认的 Sheet

    for table_info in tables:
        table_name = table_info['name']
        # Excel 工作表名称限制（不能超过31字符，不能包含特殊字符）
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', table_name)[:31]

        ws = wb.create_sheet(title=sheet_name)

        # 解析表格数据
        data = parse_table_lines(table_info['lines'])

        # 写入数据
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # 表头样式
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    print(f"Excel 文件已生成: {output_file}")
    print(f"共创建 {len(tables)} 个工作表")

if __name__ == "__main__":
    # 支持命令行参数
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        # 如果没有指定输出文件，自动生成（将 .md 替换为 .xlsx）
        if len(sys.argv) >= 3:
            output_file = sys.argv[2]
        else:
            output_file = os.path.splitext(input_file)[0] + '.xlsx'
    else:
        # 默认文件（如果没有提供参数）
        input_file = r"d:\NDC_project\配置表\Episode1_循环1_Talk表.md"
        output_file = r"d:\NDC_project\配置表\Episode1_循环1_Talk表.xlsx"

    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到输入文件: {input_file}")
        sys.exit(1)

    print(f"输入文件: {input_file}")
    print(f"输出文件: {output_file}")
    print("开始转换...")

    tables = parse_markdown_table(input_file)

    if tables:
        create_excel(tables, output_file)
        print(f"\n工作表列表:")
        for i, table in enumerate(tables, 1):
            print(f"  {i}. {table['name']}")
    else:
        print("未找到任何表格")
