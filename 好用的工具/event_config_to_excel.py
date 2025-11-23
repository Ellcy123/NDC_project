import json
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def event_config_to_excel(json_file, output_file):
    """将事件配置JSON文件转换为Excel"""

    # 读取 JSON 文件
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 检查数据结构
    if 'events' not in data:
        print("[ERROR] JSON 文件必须包含 'events' 字段")
        return False

    events = data['events']

    if len(events) == 0:
        print("[ERROR] events 数据为空")
        return False

    # 创建工作簿
    wb = openpyxl.Workbook()

    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ========== 工作表1: 事件主表 ==========
    ws_main = wb.create_sheet("事件主表")

    # 定义表头
    headers = [
        'eventId', 'eventName', 'triggerCondition',
        'unlockScenes', 'unlockDialogues', 'unlockTasks', 'unlockEvidences', 'unlockLoops',
        'unlockText_cn', 'unlockText_en',
        'ifInterrupt', 'priority'
    ]

    header_cn = [
        '事件ID', '事件名称', '触发条件',
        '解锁场景', '解锁对话', '解锁任务', '解锁证据', '解锁循环',
        '解锁提示_中文', '解锁提示_英文',
        '是否打断', '优先级'
    ]

    # 写入中文表头（第1行）
    for col_idx, header_text in enumerate(header_cn, start=1):
        cell = ws_main.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # 写入英文字段名（第2行）
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_main.cell(row=2, column=col_idx, value=header)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # 写入数据
    for row_idx, event in enumerate(events, start=3):
        # 提取unlockContent
        unlock_content = event.get('unlockContent', {})
        unlock_scenes = ", ".join(unlock_content.get('scenes', []))
        unlock_dialogues = ", ".join(unlock_content.get('dialogues', []))
        unlock_tasks = ", ".join(unlock_content.get('tasks', []))
        unlock_evidences = ", ".join(unlock_content.get('evidences', []))
        unlock_loops = ", ".join(unlock_content.get('loops', []))

        # 提取unlockText
        unlock_text = event.get('unlockText', {})
        unlock_text_cn = unlock_text.get('cn', '') if isinstance(unlock_text, dict) else unlock_text
        unlock_text_en = unlock_text.get('en', '') if isinstance(unlock_text, dict) else ''

        row_data = [
            event.get('eventId', ''),
            event.get('eventName', ''),
            event.get('triggerCondition', ''),
            unlock_scenes,
            unlock_dialogues,
            unlock_tasks,
            unlock_evidences,
            unlock_loops,
            unlock_text_cn,
            unlock_text_en,
            event.get('ifInterrupt', 0),
            event.get('priority', 50)
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # 根据优先级设置背景色
            priority = event.get('priority', 50)
            if priority >= 90:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif priority >= 70:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # 自动调整列宽
    for column in ws_main.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    cell_length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max(max_length + 2, 12), 60)
        ws_main.column_dimensions[column_letter].width = adjusted_width

    # 冻结前两行
    ws_main.freeze_panes = 'A3'

    # ========== 工作表2: 触发条件类型 ==========
    ws_condition = wb.create_sheet("触发条件类型")

    if 'conditionTypes' in data:
        condition_headers = ['条件类型', '说明']

        for col_idx, header in enumerate(condition_headers, start=1):
            cell = ws_condition.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        row = 2
        for condition_type, description in data['conditionTypes'].items():
            ws_condition.cell(row=row, column=1, value=condition_type)
            ws_condition.cell(row=row, column=2, value=description)

            for col in range(1, 3):
                cell = ws_condition.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            row += 1

        ws_condition.column_dimensions['A'].width = 25
        ws_condition.column_dimensions['B'].width = 30

    # 保存文件
    wb.save(output_file)
    print(f"[OK] Excel 文件已生成: {output_file}")
    print(f"[INFO] 共转换 {len(events)} 个事件")
    print(f"[INFO] 工作表: 事件主表, 触发条件类型")

    return True

if __name__ == "__main__":
    # 支持命令行参数
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        if len(sys.argv) >= 3:
            output_file = sys.argv[2]
        else:
            output_file = os.path.splitext(input_file)[0] + '.xlsx'
    else:
        # 默认文件
        input_file = r"d:\NDC_project\Config\EventConfig_Episode1_Loop1.json"
        output_file = r"d:\NDC_project\Config\EventConfig_Episode1_Loop1.xlsx"

    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"[ERROR] 找不到输入文件: {input_file}")
        sys.exit(1)

    print(f"[INFO] 输入文件: {input_file}")
    print(f"[INFO] 输出文件: {output_file}")
    print("[INFO] 开始转换...")
    print()

    success = event_config_to_excel(input_file, output_file)

    if not success:
        sys.exit(1)
