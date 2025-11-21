# -*- coding: utf-8 -*-
"""
JSON to Excel 转换工具 (支持三行表头: 中文名/字段名/类型)
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 字段中文名映射
FIELD_CN_NAMES = {
    "sceneId": "场景ID",
    "loopId": "循环ID",
    "sceneName": "场景名称(中文)",
    "sceneNameEn": "场景名称(英文)",
    "chapterId": "章节ID",
    "sceneType": "场景类型",
    "backgroundImage": "背景图片路径",
    "backgroundMusic": "背景音乐ID",
    "ambientSound": "环境音效ID",
    "unlockCondition": "解锁条件",
    "unlockText": "解锁提示文本",
    "npcsPresent": "场景中的NPC",
    "canSearch": "是否可搜证",
    "canLeave": "是否可离开",
    "initialDialogue": "初始对话ID",
    "description": "场景描述",
    "备注": "备注"
}

# 字段类型映射
FIELD_TYPES = {
    "sceneId": "string",
    "loopId": "string",
    "sceneName": "string",
    "sceneNameEn": "string",
    "chapterId": "string",
    "sceneType": "enum(crime/dialogue/locked)",
    "backgroundImage": "string",
    "backgroundMusic": "string",
    "ambientSound": "string",
    "unlockCondition": "string",
    "unlockText": "string",
    "npcsPresent": "string",
    "canSearch": "bool",
    "canLeave": "bool",
    "initialDialogue": "string",
    "description": "string",
    "备注": "string"
}

def json_to_excel_with_header(json_file, output_file):
    """
    将JSON文件转换为Excel,支持三行表头
    """
    try:
        # 读取JSON文件
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if not isinstance(data, list) or len(data) == 0:
            print("JSON文件必须包含非空数组")
            return False

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "SceneConfig"

        # 获取字段列表
        fields = list(data[0].keys())

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill_cn = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_fill_field = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        header_fill_type = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 写入第一行: 中文名称
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = FIELD_CN_NAMES.get(field, field)
            cell.font = header_font
            cell.fill = header_fill_cn
            cell.alignment = center_alignment

        # 写入第二行: 字段名
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = field
            cell.font = Font(size=10, italic=True)
            cell.fill = header_fill_field
            cell.alignment = center_alignment

        # 写入第三行: 字段类型
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = FIELD_TYPES.get(field, "string")
            cell.font = Font(size=9, color="666666")
            cell.fill = header_fill_type
            cell.alignment = center_alignment

        # 写入数据行
        for row_idx, item in enumerate(data, 4):  # 从第4行开始
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = item.get(field, "")

                # 处理布尔值
                if isinstance(value, bool):
                    cell.value = "TRUE" if value else "FALSE"
                else:
                    cell.value = value

                cell.alignment = left_alignment

        # 自动调整列宽
        for col_idx, field in enumerate(fields, 1):
            column_letter = get_column_letter(col_idx)

            # 计算最大宽度
            max_length = 0

            # 检查表头宽度
            header_cn = FIELD_CN_NAMES.get(field, field)
            max_length = max(max_length, len(str(header_cn)) * 2)  # 中文字符宽度
            max_length = max(max_length, len(field) * 1.2)
            max_length = max(max_length, len(FIELD_TYPES.get(field, "")) * 1.2)

            # 检查数据宽度
            for row_idx in range(4, len(data) + 4):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # 中文字符算2个宽度,英文算1个
                    length = sum(2 if ord(c) > 127 else 1 for c in str(cell_value))
                    max_length = max(max_length, length)

            # 设置列宽 (限制最大宽度)
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 冻结前3行(表头)
        ws.freeze_panes = "A4"

        # 保存文件
        wb.save(output_file)

        print(f"Excel文件已生成: {output_file}")
        print(f"已转换 {len(data)} 条数据, {len(fields)} 个字段")
        return True

    except Exception as e:
        print(f"转换失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python json_to_excel_with_header.py <输入JSON文件> [输出Excel文件]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.replace('.json', '.xlsx')

    print(f"输入文件: {input_file}")
    print(f"输出文件: {output_file}")
    print("开始转换...")

    success = json_to_excel_with_header(input_file, output_file)
    sys.exit(0 if success else 1)
