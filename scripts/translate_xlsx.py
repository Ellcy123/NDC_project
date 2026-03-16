import openpyxl
from openpyxl.styles import Font, Alignment
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('C:/Users/Ellcy/Downloads/第一集细节修改内容详情.xlsx')
ws = wb['Sheet1']

col_idx = 8
ws.cell(row=1, column=col_idx, value='程序描述（翻译）')
ws.cell(row=1, column=col_idx).font = Font(bold=True)

translations = {}

translations[3] = (
    "【场景系统 - 位置记忆】从场景A进入场景B后返回A时，需恢复离开A时的Camera位置"
    "（ScrollRect偏移量/Camera Transform）。\n"
    "实现：SceneManager中缓存每个场景的相机状态，返回时读取恢复。"
)

translations[4] = (
    "【场景系统 - Bug】部分场景加载时画面有一帧跳动。\n"
    "疑似场景初始化时Camera/RectTransform位置在第一帧未设置到正确坐标。\n"
    "排查场景加载时Camera初始化时序。"
)

translations[5] = (
    "【场景切换 - 转场动画】点击门/出入口热点时：\n"
    "Camera向门方向DOTween位移+缩放(ZoomIn) → CanvasGroup.DOFade(黑) → LoadScene → FadeIn。"
)

translations[6] = (
    "【场景探索 - NPC交互提示】鼠标悬停场景NPC时，显示对话Icon，200ms FadeIn；\n"
    "鼠标移出时反向FadeOut。用PointerEnter/PointerExit触发CanvasGroup.alpha动画。"
)

translations[7] = (
    "【AVG对话 - 镜头与布局】进入NPC对话时：\n"
    "1) Camera向NPC方向微推进（Dolly In）；\n"
    "2) 背景高斯模糊+降亮度（Post Processing或Overlay遮罩）；\n"
    "3) 策划需确定AVG角色立绘的固定屏幕坐标（左/右说话人X,Y）。"
)

translations[8] = (
    "【UI - 菜单可见性】左上角主菜单仅在场景探索状态下显示，"
    "进入AVG对话/Expose指证时隐藏。添加Slide/Fade展开收起过渡动画。"
)

translations[9] = (
    "【UI - 疑点系统】Doubt面板纸张纹理背景Sprite与设计稿不一致，需替换。"
)

translations[10] = (
    "【音频 - SFX资源清单】需导入以下音效：\n"
    "UI点击、鼠标悬停、开门、关门、脚步声、上楼梯、道具入包、\n"
    "拍照快门、疑点解锁、证词提取、背包悬停。"
)

translations[11] = (
    "【音频 - SFX绑定】将音效AudioClip绑定到对应UI事件和游戏事件上"
    "（AudioSource.PlayOneShot），逐个确认触发点。"
)

translations[12] = (
    "【UI - 疑点解锁动效】Doubt首次解锁时的弹出动画未实现。参考sample视频。"
)

translations[13] = (
    "【UI - 调查面板Tab】INQUIRY面板中缺少\"嫌疑人(Suspects)\"和\"OTHER\"两个Tab页签UI。"
)

translations[14] = (
    "【场景探索 - Bug/配置】部分可交互热点在收集信息后没有显示红圈（已探索标记）。\n"
    "可能原因：1) 已探索标记的条件判断逻辑有误；2) 配置表中该热点缺少标记配置。"
)

translations[15] = (
    "【UI - 层级管理】在子场景弹窗中点击物品打开详情面板时，"
    "需自动关闭底层的子场景面板，避免UI层叠。"
)

translations[16] = (
    "【UI - 证据Hover交互】所有证据/道具图标的Hover状态：\n"
    "1) 高亮边框(Outline Shader或Image边框)；\n"
    "2) 鼠标指针变放大镜Cursor（Cursor.SetCursor）；\n"
    "3) 图标轻微旋转(DORotate +/-3度)或位移(DOMove 2-3px)；\n"
    "4) 播放悬停音效。"
)

translations[17] = (
    "【逻辑 - NPC状态指示器】NPC感叹号\"!\"显示规则不完整：\n"
    "当前证词提取完成后仍显示\"!\"。需实现状态机：\n"
    "HasNewDialogue=显示! → AllExtracted=隐藏! → 新循环有新内容=重新显示!。"
)

translations[18] = (
    "【UI - Bug】NOTE面板中点击NPC头像进入角色详情时，"
    "位置有跳动（RectTransform初始值或Layout重建导致一帧位移）。"
)

translations[19] = "【UI - 暂停菜单】制作ESC键暂停/设置菜单（继续游戏、设置、退出等）。"

translations[20] = (
    "【UI - 证词显示优化】\n"
    "1) 调整Text.lineSpacing行距；\n"
    "2) 无证词或自述时隐藏Section标题（list.Count==0 → SetActive(false)）；\n"
    "3) 提取时添加证词文本飞向提取区域的DOTween位移动画。"
)

translations[21] = (
    "【UI - 时间线样式】Timeline面板的Font/FontSize和lineSpacing与设计稿不一致，"
    "对照设计稿调整Text组件参数。"
)

translations[22] = "【策划确认】关系网各角色头像初始坐标需策划确定后配置。"

translations[23] = (
    "【UI - 关系网】关系网面板中角色连线样式（LineRenderer或UI Line）"
    "应与OVERVIEW面板使用相同线条样式（颜色、粗细、虚实）。"
)

translations[25] = (
    "【UI - 指证成功效果】Expose成功时：\n"
    "1) 盖章(Stamp)动画——图章从上方落下+Scale回弹；\n"
    "2) 盖章背后烟雾粒子特效(Particle System)。"
)

translations[26] = (
    "【场景氛围 - 通用粒子特效】所有场景添加环境氛围效果：\n"
    "灯光闪烁(Light.intensity动画)、烟雾(Particle System)、空气尘埃粒子。\n"
    "做成通用Prefab，可按场景配置开关。"
)

translations[27] = (
    "【视频播放】播放VideoPlayer期间隐藏鼠标（Cursor.visible=false），"
    "播放结束后恢复。"
)

translations[28] = (
    "【场景探索 - 点击反馈】点击空白区域时：\n"
    "1) 点击位置生成圆形涟漪(Circle Sprite + DOScale + DOFade)；\n"
    "2) 放大镜光标Scale脉冲（1.0→1.2→1.0）。"
)

translations[29] = (
    "【场景探索 - 拾取效果】\n"
    "1) 物品图标带Ease的淡入淡出；\n"
    "2) 拾取后物品沿曲线(DOPath)飞向菜单栏背包图标。"
)

translations[30] = (
    "【UI - 证词提取页】证词提取界面的\"地点\"筛选Dropdown未实现。\n"
    "[疑问：\"地址下拉菜单\"是指按场景/地点筛选证词列表的功能吗？请确认。]"
)

translations[31] = (
    "【UI - 菜单通知气泡】菜单栏按钮添加红色数字Badge通知，\n"
    "出现时弹跳动画(DOScale: 0→1.2→1.0, EaseOutBack)。\n"
    "有新证据/疑点/证词更新时显示。"
)

translations[32] = (
    "【UI - Alert提示条】全局Toast提示修正：\n"
    "文字FontSize调大；添加半透明背景色块Image作为底板。\n"
    "（注：Alert指游戏内的操作提示/系统提示弹出条）"
)

translations[33] = (
    "【视频 - 开场CG重剪】\n"
    "1) 开头增加编剧署名和立意文本字幕叠加；\n"
    "2) 整体节奏放慢（重剪或调整播放速率）。"
)

translations[34] = (
    "【策划确认】字幕系统技术方案待定：数据格式、时间轴对齐规则、多语言支持、输出格式。"
)

translations[35] = (
    "【UI - 标题界面】开场报纸画面上的文字改为\"开始游戏\"，字号缩小2号。\n"
    "[策划确认：标题格式是\"第一章 - xxxx\"还是其他？]"
)

translations[36] = (
    "【UI - 图标对齐】\"提取完成\"和\"分析完成\"的状态小Icon锚点/位置有偏移，需重新对齐。"
)

translations[37] = (
    "【配置 - 场景入口图标】部分场景进/出门热点使用了错误Icon Sprite"
    "（门图标错显为楼梯图标）。检查SceneConfig中热点的icon字段。"
)

translations[38] = (
    "【UI - 场景名称标签】每个场景入口/出口热点Icon旁添加场景名称Text标签。"
)

translations[39] = (
    "【功能 - 新手引导】制作Tutorial系统，策划补充详细流程。\n"
    "初步：MAP按钮在第一次需要时才显示（延迟展示）。"
)

translations[40] = (
    "【美术资源】制作6张\"破绽(Flaw)\"插图，对应Loop1-6各一张。"
)

translations[41] = (
    "【功能 - 破绽展示】Expose指证中\"破绽\"功能：\n"
    "仅在每个NPC的最后一轮终极指证(Final Round)时出现。\n"
    "需确定表现形式（图片展示+动画）。"
)

translations[42] = (
    "【功能 - 电话事件】制作突发事件中的电话呼叫交互环节。\n"
    "[疑问：具体是哪个剧情事件？Loop几的什么电话？需策划提供流程。]"
)

translations[43] = "【小游戏】\"敲泥罐\"小游戏。具体玩法需求待策划补充。"
translations[44] = "【小游戏】\"拼船票\"小游戏（碎片拼图类？）。具体需求待策划补充。"
translations[45] = "【小游戏】Rosa储物室箱子解谜小游戏。具体需求待策划补充。"
translations[46] = "【小游戏】\"拼凹痕\"对比匹配小游戏。具体需求待策划补充。"

translations[47] = (
    "【UI - Bug】Expose指证结束后的对话画面整体偏左未居中。\n"
    "检查对话UI的RectTransform锚点或父容器Anchor/Pivot设置。"
)

translations[48] = (
    "【UI - CASEBOARD动效】案件板面板打开时添加展开动画，配合已有音效同步。"
)

translations[49] = (
    "【UI - 按钮Hover动效】鼠标悬停\"调查\"按钮时添加Hover动画（高亮/缩放/颜色变化）。"
)

translations[50] = (
    "【UI - 调查面板布局】Inquiry Panel所有UI元素位置与设计稿不一致，\n"
    "逐项修正：lineSpacing、spacing/padding、顶部NPC头像坐标。"
)

translations[51] = (
    "【策划确认 - 循环结算流程】Loop完成后流程：\n"
    "1) Expose结束 → 2) Zack与Emma复盘对话 → \n"
    "3) UI弹窗显示本轮攻破内容，玩家点击\"更新\"后关闭 → \n"
    "4) OVERVIEW菜单亮起红点。\n"
    "是否需要额外的\"循环总结\"界面？"
)

translations[52] = "【UI - 按钮Hover动效】\"展开调查\"按钮的Hover动画（同Row49类似）。"

translations[53] = "【视频 - 重剪】结尾真相揭示视频需重新剪辑。"

translations[54] = (
    "【UI/动效 - 指证打击感】Expose视觉冲击力优化：\n"
    "1) 对照sample视频还原打击节奏（屏幕震动/闪白Screen Shake/Flash）；\n"
    "2) 指证面板添加噪点(Noise Grain)闪烁；\n"
    "3) UI中增加Zack头像；\n"
    "4) 指证状态下UI位置还原到设计稿；\n"
    "5) 背景缓动+面板打开节奏感。"
)

translations[55] = (
    "【策划确认 - 指证成功视频】方案待定：\n"
    "A=所有指证共用统一过场视频；\n"
    "B=根据层级(普通/关键/终极)分别制作不同冲击力视频。"
)

translations[56] = (
    "【场景探索 - 平移手感】场景左右Pan/Scroll响应速度加快。\n"
    "调整ScrollRect.scrollSensitivity或Camera移动速度参数。"
)

translations[57] = (
    "【AVG对话 - 道具弹窗】AVG对话中script:\"get\"触发获取道具时，\n"
    "弹出道具详情窗口供玩家查看。实现弹窗显示/关闭逻辑。"
)

translations[58] = (
    "【UI - Bug】打字机风格Map地图UI位置不正确，需调整RectTransform坐标。"
)

translations[59] = (
    "【音频 - Bug】部分NPC语音文件有刺耳噪音/爆音(Clipping)，\n"
    "需定位具体音频文件并重新处理。"
)

translations[60] = "【QA】对所有AVG视频内容做全面检查，确认正确性和播放效果。"

translations[61] = (
    "【UI - 字体】标题画面底部COPYRIGHT文字的Font Asset与设计稿不一致，需替换。"
)

translations[62] = (
    "【UI - AVG名称对齐】对话中角色名称对齐规则：\n"
    "右侧说话人 → 名称左边缘对齐视频左边缘；\n"
    "左侧说话人 → 名称右边缘对齐视频右边缘。\n"
    "调整Text的Anchor和Alignment。"
)

translations[63] = (
    "【UI - 存档状态】无存档时\"继续游戏\"按钮应完全隐藏(SetActive(false))，\n"
    "而非灰显(interactable=false)。"
)

translations[64] = (
    "【UI - 道具查看布局】\n"
    "1) 拾起道具/线索详情界面隐藏右上角BACK按钮；\n"
    "2) 环境叙事弹窗CLOSE按钮移至屏幕下方标准位置；\n"
    "3) 图片/标题/内容的RectTransform坐标重新确认。"
)

translations[65] = (
    "【UI - 道具操作按钮】\n"
    "1) 修复背景Bar与左侧Icon重叠的层级/位置问题；\n"
    "2) 重新设计\"分析\"\"拾起\"\"拍照\"三个按钮样式。"
)

translations[66] = (
    "【UI - 子菜单过渡】二级菜单打开300ms CanvasGroup.DOFade淡入，"
    "关闭200ms淡出。"
)

translations[67] = (
    "【UI - 主菜单调整】\n"
    "1) 排序：TAB→CASEBOARD→INQUIRY→MAP→OVERVIEW→AWARDS；\n"
    "2) 文字全大写+纯黑背景框Image；\n"
    "3) Hover时Icon轻微DORotate倾斜或DOMove位移；\n"
    "4) 添加Hover和Click音效。"
)

translations[68] = (
    "【功能 - 疑点收集完整流程】\n"
    "1) 疑点纸条出现 → 2) 显示\"收集\"按钮 → \n"
    "3) 点击后纸条DOPath飞向右上角疑点面板 → \n"
    "4) 面板展开，百分比DOCounter更新动画+音效 → \n"
    "5) 1秒后面板自动收起(DOFade/DOScale)。"
)

translations[69] = (
    "【UI - 疑点面板响应】Doubt面板Hover展开/收起响应优化：\n"
    "移除延迟delay，OnPointerEnter立即展开、OnPointerExit立即收起。"
)

translations[70] = (
    "【场景切换 - 返回动画】返回父场景时画面从105% Scale缓出到100%，\n"
    "时长1-1.5秒。DOScale(1.05f→1.0f, duration, Ease.OutCubic)。"
)

# Write translations
for row_num, text in translations.items():
    cell = ws.cell(row=row_num, column=col_idx, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical='top')

# Adjust column width
ws.column_dimensions['H'].width = 80

output_path = 'C:/Users/Ellcy/Downloads/第一集细节修改内容详情_程序版.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
print(f'Total translations: {len(translations)}')
