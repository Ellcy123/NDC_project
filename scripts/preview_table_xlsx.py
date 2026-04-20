"""Preview table JSON <-> Excel round-trip tool.

Converts all *.json tables under `preview_new2/data/table/` into a single
workbook that策划 can edit directly in Excel, then writes the edited
workbook back into the JSON files.

Design follows the D:\\NDC\\res\\xls convention: every table is a single
sheet with fully-flattened columns — no sub-sheets, no JSON in cells. All
arrays and nested objects are expanded into path-based column names so
each leaf value lives in its own cell.

Column naming:
  * Dict key:        `key`
  * Array index:     numeric token
  * Paths joined by `__`:
      doubts__0__id                  # doubts[0].id
      doubts__0__condition__1__type  # doubts[0].condition[1].type
      Name__0 / Name__1              # Name[0], Name[1]  (e.g. CN / EN)
  * Max width per array is set by the longest observed occurrence. Rows
    that have fewer elements leave trailing positions blank.

Cells are force-formatted as TEXT so Excel won't auto-cast "0001" → 1.

Round-trip rules:
  * Row order is preserved. Leaf column order is first-appearance across
    all rows.
  * On write-back, cell values at a path recreate the nested structure.
  * Empty cells are treated as "leaf absent" — sparse arrays are
    compacted so gaps collapse. Pass `--keep-empty-strings` to keep
    every declared leaf as an empty string.
  * Output JSON is indent=2, UTF-8, trailing newline. Two source files
    (GameFlowConfig.json and Task.json) are already malformed and will
    be skipped with a warning.

Usage:
  python scripts/preview_table_xlsx.py to-xlsx
  python scripts/preview_table_xlsx.py to-json
  python scripts/preview_table_xlsx.py to-xlsx --xlsx my.xlsx --table-dir path/to/table
  python scripts/preview_table_xlsx.py to-json --xlsx my.xlsx --table-dir path/to/table
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_TABLE_DIR = REPO_ROOT / "preview_new2" / "data" / "table"
DEFAULT_XLSX = DEFAULT_TABLE_DIR / "_all_tables.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
DATA_ALIGN = Alignment(wrap_text=True, vertical="top")


def _scrub(text: str) -> str:
    """Remove control chars openpyxl refuses. These are always data bugs."""
    return ILLEGAL_CHARACTERS_RE.sub("", text)


def _scalar_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return _scrub(str(value))


def _text_to_scalar(text):
    """Inverse of _scalar_to_text. Empty → None."""
    if text is None:
        return None
    if isinstance(text, (int, float)):
        text = str(text)
    if not isinstance(text, str):
        text = str(text)
    if text == "":
        return None
    return text


# ---------------------------------------------------------------------------
# Path planning — walk the records, emit one leaf path per unique shape
# ---------------------------------------------------------------------------


def _shape_paths(values):
    """Given a homogeneous value-slot, return relative leaf paths with a
    fallback flag: `[(path, is_fallback), ...]`.

    All array positions share one unified shape so that
    `Parameters__0__ParameterInt / Parameters__0__ParameterStr /
     Parameters__1__ParameterInt / ...` stay in consistent order.

    When the slot mixes structured (dict/list) and scalar values, the
    majority structural shape is used AND a scalar-fallback leaf is
    appended — populated only on rows whose value doesn't match the
    majority kind (e.g. Talk.Speaker = "_RECORDER" on 3 of 3089 rows).
    """
    non_empty = [v for v in values if v is not None]
    if not non_empty:
        return [((), False)]

    has_dict = any(isinstance(v, dict) for v in non_empty)
    has_list = any(isinstance(v, list) for v in non_empty)
    has_scalar = any(not isinstance(v, (dict, list)) for v in non_empty)

    if has_list and not has_dict:
        list_values = [v for v in non_empty if isinstance(v, list)]
        max_len = max(len(v) for v in list_values) if list_values else 0
        if max_len == 0:
            return [((), False)]
        all_elements = [e for v in list_values for e in v]
        sub = _shape_paths(all_elements)
        result = [((("index", i),) + sp, fb) for i in range(max_len) for sp, fb in sub]
        if has_scalar:
            result.append(((), True))
        return result

    if has_dict and not has_list:
        dict_values = [v for v in non_empty if isinstance(v, dict)]
        ordered: list[str] = []
        seen_k: set[str] = set()
        for v in dict_values:
            for k in v:
                if k not in seen_k:
                    seen_k.add(k)
                    ordered.append(k)
        if not ordered:
            return [((), False)]
        result = []
        for k in ordered:
            sub = [v[k] for v in dict_values if k in v]
            for sp, fb in _shape_paths(sub):
                result.append(((("key", k),) + sp, fb))
        if has_scalar:
            result.append(((), True))
        return result

    if has_dict and has_list:
        # Rare truly-mixed dict/list case: just emit one leaf.
        return [((), False)]

    return [((), False)]


def _plan_paths(rows):
    """Return an ordered list of (path, is_fallback) pairs."""
    top_keys: list[str] = []
    seen_top: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        for k in row:
            if k not in seen_top:
                seen_top.add(k)
                top_keys.append(k)

    paths: list[tuple] = []
    for k in top_keys:
        values = [row[k] for row in rows if k in row]
        for rel, fb in _shape_paths(values):
            paths.append(((("key", k),) + rel, fb))
    return paths


def _path_to_name(path) -> str:
    return "__".join(str(step[1]) for step in path)


_INDEX_RE = re.compile(r"^\d+$")


def _name_to_path(name: str) -> tuple:
    """Parse a flattened column name back into a path.

    First token is always a dict key. Subsequent all-digit tokens are
    interpreted as indices; other tokens as dict keys.
    """
    tokens = name.split("__")
    if not tokens:
        return ()
    path = [("key", tokens[0])]
    for tok in tokens[1:]:
        if _INDEX_RE.match(tok):
            path.append(("index", int(tok)))
        else:
            path.append(("key", tok))
    return tuple(path)


def _get_by_path(row, path):
    cur = row
    for kind, val in path:
        if cur is None:
            return None
        if kind == "key":
            if not isinstance(cur, dict):
                return None
            cur = cur.get(val)
        else:
            if not isinstance(cur, list) or val >= len(cur):
                return None
            cur = cur[val]
    return cur


def _set_by_path(root, path, value):
    cur = root
    for i, (kind, val) in enumerate(path):
        is_last = i == len(path) - 1
        next_kind = path[i + 1][0] if not is_last else None
        if kind == "key":
            if not isinstance(cur, dict):
                return
            if is_last:
                cur[val] = value
                return
            existing = cur.get(val)
            if next_kind == "index":
                if not isinstance(existing, list):
                    existing = []
                    cur[val] = existing
            else:
                if not isinstance(existing, dict):
                    existing = {}
                    cur[val] = existing
            cur = existing
        else:
            if not isinstance(cur, list):
                return
            while len(cur) <= val:
                cur.append(None)
            if is_last:
                cur[val] = value
                return
            if next_kind == "index":
                if not isinstance(cur[val], list):
                    cur[val] = []
            else:
                if not isinstance(cur[val], dict):
                    cur[val] = {}
            cur = cur[val]


def _compact(value, keep_empty=False):
    """Remove None / {} / [] entries from lists; drop null values from dicts."""
    if isinstance(value, list):
        out = []
        for v in value:
            v2 = _compact(v, keep_empty)
            if v2 is None:
                continue
            if isinstance(v2, (dict, list)) and not v2 and not keep_empty:
                continue
            out.append(v2)
        return out
    if isinstance(value, dict):
        out = {}
        for k, v in value.items():
            v2 = _compact(v, keep_empty)
            if v2 is None:
                continue
            if isinstance(v2, (dict, list)) and not v2 and not keep_empty:
                continue
            out[k] = v2
        return out
    return value


# ---------------------------------------------------------------------------
# to-xlsx
# ---------------------------------------------------------------------------


def _column_width_from_header(name: str) -> int:
    # Scale width roughly to header length; cap to avoid over-wide columns.
    base = len(name)
    return min(40, max(12, base + 2))


def _apply_text_format(ws, max_row, max_col):
    for col_idx in range(1, max_col + 1):
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = "@"
            cell.alignment = DATA_ALIGN
        header_name = ws.cell(row=1, column=col_idx).value or ""
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _column_width_from_header(str(header_name))
        )


def _write_sheet(wb, sheet_name, rows):
    paths = _plan_paths(rows)
    header = [_path_to_name(p) for p, _ in paths] if paths else ["_empty"]

    ws = wb.create_sheet(title=sheet_name[:31])
    ws.append(header)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    scrubbed = 0
    for row in rows:
        cells = []
        for path, is_fallback in paths:
            val = _get_by_path(row, path)
            if val is None:
                cells.append("")
                continue
            if is_fallback:
                # Only populated when the row's value isn't the dominant
                # structural kind; sub-columns carry structural data.
                if isinstance(val, (dict, list)):
                    cells.append("")
                else:
                    text = _scalar_to_text(val)
                    if isinstance(val, str) and text != val:
                        scrubbed += 1
                    cells.append(text)
                continue
            if isinstance(val, (dict, list)):
                text = _scrub(json.dumps(val, ensure_ascii=False))
                cells.append(text)
            else:
                text = _scalar_to_text(val)
                if isinstance(val, str) and text != val:
                    scrubbed += 1
                cells.append(text)
        ws.append(cells)

    _apply_text_format(ws, ws.max_row, ws.max_column)
    ws.freeze_panes = "A2"
    return len(paths), scrubbed


def to_xlsx(table_dir: Path, xlsx_path: Path) -> None:
    json_files = sorted(p for p in table_dir.glob("*.json"))
    if not json_files:
        print(f"[!] no *.json files under {table_dir}", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)

    written = 0
    skipped: list[tuple[str, str]] = []
    for path in json_files:
        try:
            with path.open(encoding="utf-8") as fp:
                data = json.load(fp)
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            skipped.append((path.name, str(exc)))
            continue
        if not isinstance(data, list):
            skipped.append((path.name, f"top-level is {type(data).__name__}"))
            continue

        try:
            cols, scrubbed = _write_sheet(wb, path.stem, data)
        except Exception as exc:
            skipped.append((path.name, f"{type(exc).__name__}: {exc}"))
            continue

        note = f"    [info] {path.name}: {cols} columns, {len(data)} rows"
        if scrubbed:
            note += f"  (scrubbed {scrubbed} illegal control chars)"
        print(note)
        written += 1

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"[OK] wrote {written} sheets to {xlsx_path}")
    if skipped:
        print(f"[!] skipped {len(skipped)} file(s):")
        for name, reason in skipped:
            print(f"    - {name}: {reason}")


# ---------------------------------------------------------------------------
# to-json
# ---------------------------------------------------------------------------


def _sheet_to_rows(ws, keep_empty_strings: bool = False):
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    if header is None:
        return []
    paths: list = []
    for name in header:
        if name is None:
            paths.append(None)
        else:
            paths.append(_name_to_path(str(name)))

    # A column header whose path is a prefix of another column's path is
    # a scalar-fallback column (e.g. "Speaker" alongside "Speaker__id" /
    # "Speaker__Name__0" etc). Those are applied last so structural
    # sub-columns take precedence, mirroring the writer.
    prefix_set: set[tuple] = set()
    for p in paths:
        if p is None:
            continue
        for i in range(1, len(p)):
            prefix_set.add(p[:i])
    is_fallback = [
        (p is not None and p in prefix_set) for p in paths
    ]

    rows: list[dict] = []
    for raw in it:
        if raw is None:
            continue
        if all(v in (None, "") for v in raw):
            continue
        root: dict = {}
        # Pass 1: structural sub-paths
        for i, path in enumerate(paths):
            if path is None or path == () or is_fallback[i]:
                continue
            val = raw[i] if i < len(raw) else None
            parsed = _parse_cell(val)
            if parsed is None:
                if keep_empty_strings:
                    _set_by_path(root, path, "")
                continue
            _set_by_path(root, path, parsed)
        # Pass 2: fallback paths — only apply if sub-structure is empty
        for i, path in enumerate(paths):
            if path is None or path == () or not is_fallback[i]:
                continue
            val = raw[i] if i < len(raw) else None
            parsed = _parse_cell(val)
            if parsed is None:
                continue
            existing = _get_by_path(root, path)
            if isinstance(existing, (dict, list)) and existing:
                continue
            _set_by_path(root, path, parsed)
        cleaned = _compact(root, keep_empty=keep_empty_strings)
        if cleaned:
            rows.append(cleaned)
    return rows


def _parse_cell(val):
    """Parse a cell: scalar text, or JSON if it looks like JSON."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        val = str(val)
    if not isinstance(val, str):
        val = str(val)
    if val == "":
        return None
    stripped = val.lstrip()
    if stripped.startswith(("[", "{")):
        try:
            parsed = json.loads(val)
            if isinstance(parsed, (list, dict)):
                return parsed
        except json.JSONDecodeError:
            pass
    return val


def to_json(table_dir: Path, xlsx_path: Path, keep_empty_strings: bool = False) -> None:
    if not xlsx_path.exists():
        print(f"[!] xlsx not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    table_dir.mkdir(parents=True, exist_ok=True)

    written = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _sheet_to_rows(ws, keep_empty_strings=keep_empty_strings)
        out_path = table_dir / f"{sheet_name}.json"
        with out_path.open("w", encoding="utf-8") as fp:
            json.dump(rows, fp, ensure_ascii=False, indent=2)
            fp.write("\n")
        print(f"[OK] {sheet_name} -> {out_path.name} ({len(rows)} rows)")
        written += 1
    print(f"[done] wrote {written} json file(s) under {table_dir}")


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("mode", choices=["to-xlsx", "to-json"], help="conversion direction")
    parser.add_argument(
        "--table-dir",
        type=Path,
        default=DEFAULT_TABLE_DIR,
        help=f"json directory (default: {DEFAULT_TABLE_DIR})",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"workbook path (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--keep-empty-strings",
        action="store_true",
        help="to-json only: emit empty-string values for blank cells (default: drop)",
    )
    args = parser.parse_args()

    if args.mode == "to-xlsx":
        to_xlsx(args.table_dir, args.xlsx)
    else:
        to_json(args.table_dir, args.xlsx, keep_empty_strings=args.keep_empty_strings)


if __name__ == "__main__":
    main()
