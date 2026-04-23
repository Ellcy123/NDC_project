"""
配表小助手 - 数据转换脚本
将 Preview/data 数据转换为 Luban 配置表格式
"""
import sys
import io
# 修复Windows终端编码问题
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import yaml
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
import shutil
from datetime import datetime

# 路径配置
PREVIEW_DATA = Path(r"D:\NDC_project\Preview\data")
STORY_OUTPUT = Path(r"D:\NDC_project\story")
UNITY_OUTPUT = Path(r"D:\NDC\Config\Datas\story")

# 确保yaml输出中文不转义
yaml.add_representer(str, lambda dumper, data: dumper.represent_scalar('tag:yaml.org,2002:str', data))


class ConfigTableConverter:
    """配置表转换器"""

    # NPC ID 到角色编号的映射
    # 注意：角色编号用于生成 talk_id，格式 NNXXYYY
    NPC_NUMBER_MAP = {
        'NPC101': 1,   # 查克 (Zack Brennan) - 主角
        'NPC102': 2,   # 艾玛 (Emma O'Malley)
        'NPC103': 3,   # 罗莎 (Rosa Martinez)
        'NPC104': 4,   # 莫里森警探 (Detective Morrison)
        'NPC105': 5,   # 汤米 (Tommy)
        'NPC106': 6,   # 薇薇安 (Vivian)
        'NPC107': 7,   # 韦伯 (Webb)
        'NPC108': 8,   # 安娜 (Anna Webb)
        'NPC109': 9,   # 詹姆斯 (James)
        'NPC110': 10,  # 莫里森夫人 (Mrs. Morrison)
        'NPC111': 11,  # Anna (James's wife)
    }

    def __init__(self):
        self.npcs = {}
        self.scenes = {}
        self.evidences = {}
        self.loops = {}
        self.dialogs = {}
        # 记录每个NPC在每个循环的段落计数器
        self.npc_segment_counter = {}

    def load_all_data(self, unit: str = "Unit1"):
        """加载所有数据"""
        print("📖 加载数据...")

        # 加载主数据 (Unit1/master/ 目录)
        unit_path = PREVIEW_DATA / unit
        self.npcs = self._load_yaml(unit_path / "master" / "npcs.yaml")
        self.scenes = self._load_yaml(unit_path / "master" / "scenes.yaml")
        self.evidences = self._load_yaml(unit_path / "master" / "evidences.yaml")

        # 加载循环数据
        loops_dir = unit_path / "loops"
        if loops_dir.exists():
            for f in loops_dir.glob("loop*.yaml"):
                loop_num = f.stem.replace("loop", "")
                self.loops[loop_num] = self._load_yaml(f)

        # 加载对话数据
        dialogs_dir = unit_path / "dialogs"
        if dialogs_dir.exists():
            for loop_dir in dialogs_dir.iterdir():
                if loop_dir.is_dir() and loop_dir.name.startswith("loop"):
                    loop_num = loop_dir.name.replace("loop", "")
                    self.dialogs[loop_num] = {}
                    for f in loop_dir.glob("*.yaml"):
                        self.dialogs[loop_num][f.stem] = self._load_yaml(f)

        print(f"  ✅ NPCs: {len(self.npcs.get('npcs', {}))}")
        print(f"  ✅ Scenes: {len(self.scenes.get('scenes', {}))}")
        print(f"  ✅ Evidences: {len(self.evidences.get('evidences', {}))}")
        print(f"  ✅ Loops: {len(self.loops)}")
        print(f"  ✅ Dialog files: {sum(len(d) for d in self.dialogs.values())}")

    # ==================== ID 生成辅助方法 ====================

    def _get_npc_number(self, npc_id: str) -> int:
        """将 NPC ID 转换为角色编号 (NPC101 -> 1)"""
        return self.NPC_NUMBER_MAP.get(npc_id, 0)

    def _get_dialog_npc(self, dialog_data: dict, dialog_name: str) -> str:
        """获取对话文件的主 NPC ID"""
        if not dialog_data:
            return ''

        # 1. npc_dialog 类型：使用 npc 字段
        if dialog_data.get('npc'):
            return dialog_data['npc']

        # 2. accusation 类型：使用 target 字段
        if dialog_data.get('target'):
            return dialog_data['target']

        # 3. opening/ending 类型：从对话内容找出现最多的 NPC
        npc_counts = {}
        for section_key, section in dialog_data.items():
            if not isinstance(section, dict) or 'lines' not in section:
                continue
            for line in section.get('lines', []):
                speaker = line.get('speaker', '')
                if speaker.startswith('NPC') and speaker != 'NPC101':  # 排除主角
                    npc_counts[speaker] = npc_counts.get(speaker, 0) + 1

        if npc_counts:
            return max(npc_counts, key=npc_counts.get)

        return ''

    def _get_next_segment(self, loop_num: str, npc_id: str) -> int:
        """获取指定 NPC 的全局下一个段落号（跨循环累加）"""
        # 使用 NPC ID 作为 key，不包含 loop_num，这样段落号跨循环递增
        if npc_id not in self.npc_segment_counter:
            self.npc_segment_counter[npc_id] = 0
        self.npc_segment_counter[npc_id] += 1
        return self.npc_segment_counter[npc_id]

    def _sort_dialog_files(self, loop_dialogs: dict) -> List[tuple]:
        """按处理顺序排序对话文件: opening -> npc_dialogs(字母序) -> accusation -> ending"""
        sorted_files = []

        # 1. opening 优先
        if 'opening' in loop_dialogs:
            sorted_files.append(('opening', loop_dialogs['opening']))

        # 2. npc_dialog 文件（按文件名字母序）
        npc_files = []
        for name, data in loop_dialogs.items():
            if name in ['opening', 'accusation', 'ending', 'schema_dialogs']:
                continue
            if data and data.get('type') == 'npc_dialog':
                npc_files.append((name, data))
            elif data and data.get('npc'):  # 有 npc 字段的也算
                npc_files.append((name, data))
            elif name not in ['opening', 'accusation', 'ending'] and data:
                # 其他文件也按 npc_dialog 处理
                npc_files.append((name, data))

        # 按文件名字母序排序
        npc_files.sort(key=lambda x: x[0])
        sorted_files.extend(npc_files)

        # 3. accusation
        if 'accusation' in loop_dialogs:
            sorted_files.append(('accusation', loop_dialogs['accusation']))

        # 4. ending 最后
        if 'ending' in loop_dialogs:
            sorted_files.append(('ending', loop_dialogs['ending']))

        return sorted_files

    def _load_yaml(self, path: Path) -> dict:
        """加载yaml文件"""
        if not path.exists():
            print(f"  [WARN] 文件不存在: {path}")
            return {}
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            print(f"  [ERROR] 加载失败 {path.name}: {e}")
            return {}

    # ==================== 翻译辅助 ====================

    def translate(self, cn_text: str) -> str:
        """将中文翻译为英文（使用预定义映射）"""
        if not cn_text:
            return ''

        # 综合翻译映射
        translations = {
            # ===== 物品描述翻译 - 循环1 =====
            "通缉\"疤面Tony\"的悬赏金高达5000美元": "A wanted poster with a $5,000 bounty on 'Scarface Tony'",
            "照片背面写着\"我的小天使Miguel，生日0915，妈妈的一切希望\"": "On the back: 'My little angel Miguel, birthday 0915, mom's only hope'",
            "女儿Miguel的昂贵药物清单，总费用远超Rosa的工资收入，经济压力巨大": "Expensive medication list for son Miguel, total cost far exceeds Rosa's wages",
            "一条普通的白色毛巾": "An ordinary white towel",
            "接近闻嗅时有明显的甜腻化学味，是氯仿的味道，这不是清洁用品": "A distinct sweet chemical smell - chloroform, not cleaning supplies",
            "Rosa Martinez - 11月15日夜班：后台走廊清洁 23:00-01:00，明确显示工作地点是后台走廊，不是地下室酒窖": "Rosa Martinez - Nov 15 night shift: Backstage corridor 23:00-01:00, clearly shows work location is backstage corridor, not basement",
            "医用麻醉剂氯仿的玻璃瓶，在歌舞厅一楼走廊垃圾桶内发现，瓶口有少量氯仿残留": "Medical chloroform bottle found in first floor corridor trash, with residue at the opening",
            "地板上的拖拽痕迹": "Drag marks on the floor",
            "压痕较深，被拖动的东西至少150磅，普通女性的力量基本无法完成": "Deep marks, dragged object weighed at least 150 pounds, beyond an ordinary woman's strength",
            "Tommy在办公室整理账目时，于11点30分听到了枪声。这声枪响和平时黑帮火拼的声音不太一样，只听到了一声": "Tommy heard a gunshot at 11:30 while doing accounts. Sounded different from usual gang shootouts - just one shot",

            # ===== 物品描述翻译 - 循环2 =====
            "Morrison于00:30从家出发，声称处理紧急案件": "Morrison left home at 00:30, claiming to handle an urgent case",
            "案发前一天19:30领用便携式现场勘验箱，里面有各种测量设备，有Morrison签名": "Checked out portable crime scene kit at 19:30 the day before, contains measuring equipment, signed by Morrison",
            "便携式现场勘验箱，里面有各种测量设备，可与现场压痕进行比对": "Portable crime scene kit with measuring equipment, can be compared with scene imprints",
            "Morrison家中发现的各种收据，包含一些可疑的支出记录": "Various receipts found in Morrison's home with suspicious expense records",
            "欠疤面Tony $5000赌债，催债信件威胁家人安全": "Owes Scarface Tony $5,000 gambling debt, collection letters threatening family",
            "神秘人\"Whale\"的联系方式和简单指示，$5000已转入账户": "Mysterious 'Whale's contact info and simple instructions, $5,000 transferred to account",
            "几枚不同面额的赌场筹码，其中一枚背面用小字刻着\"最后一次\"": "Several casino chips of different denominations, one engraved 'last time' on back",
            "Miller工业集团遭神秘勒索，警方介入调查": "Miller Industrial Group extorted by mystery person, police investigating",
            "一枚\"英勇服务奖章\"，表彰Morrison在一次银行抢劫案中的英勇表现": "Medal of Valor commending Morrison for bravery in a bank robbery case",
            "一个普通的四方形压痕": "An ordinary rectangular imprint",
            "四方形的特殊压痕，完全吻合便携式现场勘验箱的尺寸": "Special rectangular imprint, perfectly matches the crime scene kit dimensions",
            "早期Vivian与Webb的幸福合影": "Early happy photo of Vivian and Webb together",
            "11:00左右Vivian去过Webb办公室，11:20去过街角杂货店，00:10回到化妆室": "Vivian visited Webb's office around 11:00, went to corner store at 11:20, returned to dressing room at 00:10",
            "Morrison从后门进入，直接去了Webb的办公室": "Morrison entered through back door, went directly to Webb's office",
            "Morrison检查现场时没有使用任何工具，只是草草看了几眼就出来了，像是早就知道现场情况": "Morrison didn't use any tools checking the scene, just glanced around briefly as if he knew the layout beforehand",

            # ===== 物品描述翻译 - 循环3 =====
            "月总收入15,000美元，与酒类销售差额高达11,000美元来源不明": "Monthly revenue $15,000, $11,000 gap from liquor sales of unknown origin",
            "其中有比较明确的当月勒索收入，被包装为古董收入": "Clear extortion income for the month, disguised as antique revenue",
            "详细记录客户访问时间、观看古董类型，同一古董多次\"销售\"记录": "Detailed records of client visits, antique viewings, same antique 'sold' multiple times",
            "经过笔迹对比分析，字迹与Tommy的字迹一致": "Handwriting analysis confirms it matches Tommy's handwriting",
            "三张照片，包括和多位芝加哥政商要人合影照片": "Three photos including shots with several Chicago political and business figures",
            "Webb写给Vivian的甜言蜜语信件，承诺带她去法国": "Webb's sweet-talking letter to Vivian, promising to take her to France",
            "给Vivian的钻石项链": "Diamond necklace for Vivian",
            "\"这些道貌岸然的家伙...他们的虚伪就是我的财富\" 下面有便签写着\"WHALE - DANGER\"": "'These hypocrites... their hypocrisy is my fortune' with a note below: 'WHALE - DANGER'",
            "Thompson议员被勒索后写的妥协信件": "Senator Thompson's compromise letter after being blackmailed",
            "Coleman银行家向Webb支付\"古董费用\"的转账记录": "Banker Coleman's transfer records for 'antique fees' to Webb",
            "Bennington陶器实际价值$2000，欧洲油画实际价值$150": "Bennington pottery actual value $2,000, European painting actual value $150",
            "月净酒类销售收入仅4,000美元，与声称的总收入不符": "Net monthly liquor sales only $4,000, doesn't match claimed total revenue",
            "作为一个有妻儿的男人，看着Vivian小姐被迫做那些不光彩的事情，我深感羞愧（可用于笔迹对比）": "As a family man, watching Miss Vivian forced to do shameful things fills me with guilt (can compare handwriting)",
            "Vivian与特定VIP客户的\"私人服务\"时间安排": "Vivian's 'private service' schedule with specific VIP clients",
            "照片中Tommy把妻儿环抱在臂弯里": "Photo of Tommy embracing his wife and children",
            "欧洲油画完好摆放原位，与\"已售给Coleman\"矛盾": "European painting still in place, contradicts 'sold to Coleman'",
            "中国花瓶完好摆放原位，与\"已售给Thompson\"矛盾": "Chinese vase still in place, contradicts 'sold to Thompson'",
            "Webb说'别想这么多，多看看我送你的书'，另一个人说'你不要接这个生意，会出大事的'": "Webb said 'Don't think too much, read the books I gave you', another person said 'Don't take this job, something bad will happen'",
            "那时候我在清洁，音乐实在太响了，震得我耳朵疼。那种音量下就算有别的声音我也绝对听不到": "I was cleaning then, the music was so loud it hurt my ears. I couldn't hear anything else at that volume",
            "James从Webb先生办公室出来，抱着一个水壶...非常小心翼翼...看到我立刻把水壶藏到身后": "James came out of Mr. Webb's office carrying a vase... very carefully... hid it behind his back when he saw me",

            # ===== 物品描述翻译 - 循环4 =====
            "看似普通的陶制水壶": "Seemingly ordinary clay pot",
            "表面人为涂抹泥浆伪装，清理后露出古董底色": "Surface deliberately smeared with mud to disguise, reveals antique underneath when cleaned",
            "James自草的推荐信，高度自我夸奖工作能力和忠诚度": "James's self-written recommendation letter, praising his own abilities and loyalty",
            "厨房专用手套，指甲缝隙和接缝处残留着极少量的褐色泥土痕迹": "Kitchen gloves with traces of brown mud in nail gaps and seams",
            "当晚有送餐记录，但11点到11点30分没有送餐记录": "Delivery records that night, but no deliveries between 11:00 and 11:30",
            "James与妻子Anna的温馨合影，写着\"我的世界\"": "Warm photo of James with wife Anna, inscribed 'My World'",
            "如果你想清楚了，可以联系我。纸张纹样与Morrison处发现的纸条相同": "If you've made up your mind, contact me. Paper pattern matches note found at Morrison's",
            "我亲爱的Anna，如果上帝召唤我回家...愿上帝宽恕我的软弱": "My dear Anna, if God calls me home... may God forgive my weakness",
            "大额现金收入记录，标注\"Whale\"和$5000金额": "Large cash income record, marked 'Whale' and $5,000",
            "Bennington陶器的照片": "Photo of Bennington pottery",
            "由证据320和证据425合并，Bennington陶器实际价值$2000，证明James已知古董价值": "Merged from evidence 320 and 425, Bennington pottery worth $2,000, proves James knew the value",
            "书页边缘写满了拼写练习，\"American Dream\"被反复练习": "Book margins filled with spelling practice, 'American Dream' written repeatedly",
            "合法身份申请表被撕成碎片，申请费$1500": "Legal status application torn to pieces, application fee $1,500",
            "WHALE很危险 便签": "'WHALE is dangerous' sticky note",
            "Tommy目击James从Webb办公室拿走陶土水壶，行为异常谨慎": "Tommy witnessed James taking clay pot from Webb's office, acting unusually careful",
            "Rosa看到James用钥匙进入Webb会客室，但那里没有值钱东西，行为可疑": "Rosa saw James use key to enter Webb's parlor, but nothing valuable there, suspicious behavior",

            # ===== 物品描述翻译 - 循环5 =====
            "Vivian写的分手信正文": "Body of Vivian's breakup letter",
            "信封上写着\"Webb收\"": "Envelope addressed 'To Webb'",
            "Vivian亲笔写的分手信，\"我已经受够了，我们之间结束了\"": "Vivian's handwritten breakup letter: 'I've had enough, we're over'",
            "法国邮轮船票被撕成碎片，船期是下个月，如果真要旅行不会撕毁": "French cruise ticket torn to pieces, departure next month - wouldn't tear it if really traveling",
            "如果有人背叛了我，我会让他遭受世间最痛苦的事情": "If someone betrays me, I'll make them suffer the worst pain imaginable",
            "记录与不同客户\"服务\"的日期和痛苦感受，\"我还要忍受多久\"": "Records of 'service' dates with different clients and painful feelings, 'How much longer must I endure'",
            "Webb的新欢Rita的照片和联系方式": "Photo and contact info of Webb's new lover Rita",
            "Vivian申请更改当晚演出时间的申请表": "Vivian's application to change her performance time that night",
            "案发当晚11点左右，Rosa看到Vivian神情紧张地往会客室走，并携带了可疑物品": "Around 11 PM that night, Rosa saw Vivian nervously walking to parlor with suspicious items",
            "Webb最近和一个叫Rita的女人走得很近，Vivian发现后非常痛苦，曾在化妆室哭泣说\"我为他做了这么多\"": "Webb got close to a woman named Rita recently. Vivian was devastated, cried in dressing room: 'I did so much for him'",

            # ===== 物品描述翻译 - 循环6 =====
            "窗外的一个孔洞": "A hole outside the window",
            "是一个比较新的枪击孔洞，证明11:30的枪声是故意制造的假象": "A relatively fresh bullet hole, proves the 11:30 gunshot was deliberately faked",
            "11:00-11:10是舞台鼓点最强烈时段，《芝加哥狂想》+《爵士风暴》，音量等级5级（最高）": "11:00-11:10 was loudest drumbeat period: 'Chicago Rhapsody' + 'Jazz Storm', volume level 5 (max)",
            "道具箱中的一把手枪": "A gun in the prop box",
            "那天晚上我看到Vivian离开舞台的时候悄悄把什么东西塞进了包里": "That night I saw Vivian secretly putting something in her bag when leaving stage",
            "11点10分的时候来到Webb的会客室门口，平时他都不锁门，但昨天他锁了门": "Came to Webb's parlor door at 11:10, he usually doesn't lock it, but it was locked that day",
            "当晚有送餐记录，但11点到11点30分没有任何送餐记录": "Delivery records that night, but no deliveries at all between 11:00 and 11:30",

            # ===== 对话翻译 - Loop1 Opening =====
            "[Zack猛然惊醒，头痛欲裂，眼神涣散]": "[Zack suddenly wakes up, head pounding, eyes unfocused]",
            "[看到手中的枪...]": "[Sees the gun in his hand...]",
            "什么...该死...": "What... damn...",
            "[看到Webb的尸体，Zack猛地甩开枪，惊恐后退]": "[Seeing Webb's body, Zack throws the gun away in horror]",
            "Webb？！不...这不对...": "Webb?! No... this isn't right...",
            "[他挣扎着站起，但头晕目眩，扶住墙]": "[He struggles to stand but feels dizzy, leaning against the wall]",
            "我敲门...然后...": "I knocked... then...",
            "[门被踹开！Morrison带警员冲入，枪口对准]": "[Door kicked open! Morrison storms in with officers, guns drawn]",
            "别动！举起手！": "Freeze! Hands up!",
            "等等...我是——": "Wait... I'm—",
            "闭嘴！你有权保持沉默！": "Shut up! You have the right to remain silent!",
            "我没有杀他！我是被——": "I didn't kill him! I was—",
            "被栽赃？都这么说。": "Framed? They all say that.",
            "[Morrison粗暴地给Zack铐上手铐]": "[Morrison roughly handcuffs Zack]",
            "等一下！": "Wait!",
            "滚出去！这是犯罪现场！": "Get out! This is a crime scene!",
            "《芝加哥先驱报》。我拍到有人从后门离开。": "Chicago Herald. I photographed someone leaving through the back door.",
            "你最好有证据，否则我连你一起抓！": "You better have evidence, or I'll arrest you too!",
            "我有证据，警探。不过更重要的是——一个称职的警探不会在看到现场后的第一反应就是'逮捕他'，而不是'封锁现场、寻找线索'。": "I have evidence, Detective. More importantly—a competent detective wouldn't immediately say 'arrest him' upon seeing the scene, instead of 'secure the scene, look for clues'.",
            "[Morrison脸色一变，死死盯着Emma]": "[Morrison's expression changes, staring hard at Emma]",
            "小姑娘，别多管闲事。": "Little girl, mind your own business.",
            "我只是记录事实。您要是现在抓人，明天《先驱报》头版就是：警方草率定案。": "I'm just recording facts. If you arrest him now, tomorrow's Herald headline: Police rushes to judgment.",
            "[Morrison和Emma对峙，空气凝固]": "[Morrison and Emma face off, the air frozen]",
            "听着，Brennan。72小时。找不到真凶，我让你在监狱里烂掉。": "Listen, Brennan. 72 hours. Don't find the real killer, I'll let you rot in prison.",
            "还有你。妨碍司法，够你受的。": "And you. Obstruction of justice will be enough to deal with you.",
            "[Morrison粗暴推开Emma，愤然离开]": "[Morrison shoves Emma aside and storms off]",
            "[Zack靠着墙，还在喘息，Emma走过来]": "[Zack leans against the wall, still gasping, as Emma approaches]",
            "你还好吗？": "Are you okay?",
            "谢了。": "Thanks.",
            "[沉默]": "[Silence]",
            "你到底拍到什么了？": "What exactly did you photograph?",
            "一个模糊的背影。不够当证据，但够吓唬Morrison。": "A blurry figure from behind. Not enough for evidence, but enough to scare Morrison.",
            "所以你在赌。": "So you're gambling.",
            "我们都在赌。": "We're all gambling.",
            "[Zack转身要走]": "[Zack turns to leave]",
            "你要去哪儿？": "Where are you going?",
            "找出谁陷害我。": "To find out who framed me.",
            "一个人？72小时？": "Alone? 72 hours?",
            "我习惯一个人。": "I'm used to being alone.",
            "Morrison明显有问题。他来得太快，太急于定你的罪。": "Morrison is clearly suspicious. He arrived too fast, too eager to pin this on you.",
            "所以？": "So?",
            "所以这不只是一起简单的谋杀。背后有更大的事。": "So this isn't just a simple murder. There's something bigger behind it.",
            "那是你的新闻。不是我的问题。": "That's your story. Not my problem.",
            "我也在调查Webb。我有线索，有渠道，有——": "I was also investigating Webb. I have leads, contacts—",
            "我不需要搭档。": "I don't need a partner.",
            "你需要不进监狱。": "You need to stay out of prison.",
            "[Zack沉默，看着Emma]": "[Zack falls silent, looking at Emma]",
            "听着，我知道你不信任我。我也不指望你信任我。": "Look, I know you don't trust me. I don't expect you to.",
            "但现在你没有选择。Morrison盯上你了，72小时根本不够。": "But now you have no choice. Morrison has you in his sights, 72 hours isn't nearly enough.",
            "...你想要什么？": "...What do you want?",
            "真相。还有一个好故事。": "The truth. And a good story.",
            "至少你诚实。": "At least you're honest.",
            "[Zack转身往酒吧走]": "[Zack turns and walks toward the bar]",
            "这是答应了？": "Is that a yes?",
            "我还有选择吗？": "Do I have a choice?",
            "那我们从哪儿开始？": "So where do we start?",
            "有人让我进去，有人打晕我，有人布置现场。": "Someone let me in, someone knocked me out, someone set up the scene.",
            "Webb的员工都看到我来了。酒保、门童、歌女、厨师...他们都知道些什么。": "Webb's employees all saw me arrive. Bartender, doorman, singer, cook... they all know something.",
            "好。我来问，你观察。": "Okay. I'll ask, you observe.",
            "我说了算。": "I call the shots.",
            "当然。你是专家。": "Of course. You're the expert.",
            "[Zack看了她一眼，继续往前走]": "[Zack glances at her and continues walking]",
            "我一定是疯了...": "I must be crazy...",
            "我们都是。": "We all are.",

            # ===== 对话翻译 - Loop1 Rosa =====
            "你是这儿的清洁工？": "You're the cleaner here?",
            "是...是的，先生。": "Y-yes, sir.",
            "Zack Brennan，私家侦探。昨晚你在哪儿工作？": "Zack Brennan, private detective. Where were you working last night?",
            "我...我一直在地下室。酒窖那边，整理酒瓶和架子...很忙的...一直在那儿...": "I... I was in the basement. In the wine cellar, organizing bottles and shelves... very busy... stayed there...",
            "整晚？": "All night?",
            "对！一直在！我什么都...什么都没看到！": "Yes! The whole time! I didn't... didn't see anything!",
            "我还没问你看到什么。": "I didn't ask what you saw yet.",
            "我是说...地下室很安静，我就...就一直在干活...什么都不知道...": "I mean... the basement was quiet, I just... kept working... don't know anything...",
            "地下室能听到一楼的枪声吗？": "Can you hear gunshots from the first floor in the basement?",
            "听...听不到！地下室很深，我什么都没听到...": "N-no! The basement is deep, I didn't hear anything...",
            "你在撒谎。": "You're lying.",
            "没有！我...我真的在地下室！先生，求您了...我还有个女儿要照顾...我只是个清洁工...": "No! I... I really was in the basement! Please sir... I have a daughter to take care of... I'm just a cleaner...",
            "你孩子多大？": "How old is your child?",
            "八...八岁...她生病了，需要钱治病...我真的什么都不知道...": "E-eight... she's sick, needs money for treatment... I really don't know anything...",
            "我不是来为难你的。但如果你撒谎，我会知道。": "I'm not here to make things hard for you. But if you're lying, I'll find out.",
            "我...我没有...": "I... I'm not...",
            "我会再来找你。到那时，希望你想好要说什么。": "I'll come back for you. By then, I hope you'll know what to say.",
            "好的...好的，先生...我会在这儿...": "Okay... okay, sir... I'll be here...",

            # ===== 对话翻译 - Loop1 Tommy =====
            "Brennan先生？这么晚了还来...有什么我能帮您的吗？": "Mr. Brennan? Coming so late... Is there something I can help you with?",
            "我需要问你几个问题，关于昨晚的事。": "I need to ask you a few questions about last night.",
            "昨晚...真是太不幸了。Webb先生是个...是个好老板。我在这儿工作五年了，从来没想过...会发生这种事...": "Last night... so tragic. Mr. Webb was a... a good boss. Worked here five years, never thought... this would happen...",
            "你昨晚在哪儿？": "Where were you last night?",
            "你听到枪声了吗？": "Did you hear the gunshot?",
            "Webb昨晚有没有提过他在等什么人？": "Did Webb mention he was expecting someone last night?",
            "在办公室！一直在办公室！月底了，得对账，您知道的...账目、排班、库存...这些都要我来处理...": "In the office! The whole time! Month-end, had to balance the books... accounts, schedules, inventory... all on me...",
            "有人能证明吗？": "Can anyone verify that?",
            "这个...我一个人在办公室。不过账本可以证明，您看，这些都是昨晚整理的...": "Well... I was alone in the office. But the ledgers can prove it, see, these were organized last night...",
            "账本可以伪造。还有别的证据吗？": "Ledgers can be forged. Any other evidence?",
            "好，我先记下。还有别的问题...": "Okay, I'll note that. Any other questions...",
            "我...我没有撒谎！真的！您可以去查监控...虽然后门那边没有摄像头，但是...但是我真的一直在办公室！": "I... I'm not lying! Really! Check the cameras... though there's none at the back door, but... I really was in the office!",
            "后门没有摄像头？": "No cameras at the back door?",
            "是...是的。Webb先生说那边不需要...有些...有些客人喜欢低调进出...": "Y-yes. Mr. Webb said it wasn't needed there... some... guests prefer to come and go discreetly...",
            "记下了。": "Noted.",
            "说说枪声的事": "Tell me about the gunshot",
            "Webb昨晚等的是什么人？": "Who was Webb expecting last night?",
            "谢谢您理解...我真的只是在做我的工作...": "Thank you for understanding... I was just doing my job...",
            "听到了...很响，吓了我一跳。我当时以为又是哪家的麻烦...": "I heard it... very loud, startled me. Thought it was someone else's trouble...",
            "几点？": "What time?",
            "11点半...应该是11点30分。我看了表。当时正在核对这笔账...": "11:30... should be 11:30. Checked my watch. Was going over these accounts...",
            "这声枪响和平时的一样吗？": "Did this gunshot sound the same as usual?",
            "听到枪声后你做了什么？": "What did you do after hearing the gunshot?",
            "不太一样...平时要是有麻烦，都是一连串的枪声，很乱。但昨晚...就一声。所以我当时还以为是什么东西掉了...": "Not quite the same... usually when there's trouble, it's a burst of gunfire. But last night... just one shot. Thought something fell...",
            "只有一声。": "Just one shot.",
            "对...就一声...": "Yes... just one...",
            "说说Webb的访客": "Tell me about Webb's visitors",
            "我问完了": "I'm done asking",
            "我...我没动。在这行混久了，知道有些事情...不该管的别管。我就...继续做我的账...": "I... I didn't move. Been in this business long enough to know... what you shouldn't mess with, don't. I just... kept doing my accounts...",
            "你的老板可能正在被杀，你继续做账？": "Your boss might be getting killed, and you kept doing accounts?",
            "我...我怎么知道是Webb先生出事了？这附近...这种事...唉...": "I... how was I to know it was Mr. Webb? Around here... these things... sigh...",
            "这个我...我真的不太清楚。Webb先生的私人会面，一般不会...不会告诉我这些...": "I... I really don't know much. Mr. Webb's private meetings, he usually doesn't... tell me about them...",
            "但你是经理，应该知道VIP包厢的情况。": "But you're the manager, you should know about the VIP rooms.",
            "是...是这样的，但Webb先生有时候会...会直接安排，不通过我。尤其是一些...一些重要的客人。": "Y-yes, but Mr. Webb sometimes... arranges things directly, bypassing me. Especially for some... important guests.",
            "什么样的重要客人？不说清楚别想走。": "What kind of important guests? Don't think you're leaving without explaining.",
            "我理解你的处境。但Webb死了，情况不一样了。": "I understand your situation. But Webb is dead, things are different now.",
            "Brennan先生...我...我真的...": "Mr. Brennan... I... I really...",
            "Tommy，我可以让警察来问你。你觉得他们会不会比我更有耐心？": "Tommy, I can have the police question you. Think they'll be more patient than me?",
            "别...别叫警察！我...我说！有些客人...是那边来的...您知道的...那边...": "Don't... don't call the police! I... I'll talk! Some guests... came from over there... you know... that side...",
            "Whale的人？": "Whale's people?",
            "我...我没说是谁！我只是...Webb先生有时候会接待一些...不方便公开的客人...就这些！我真的不知道更多了！": "I... I didn't say who! I just... Mr. Webb sometimes hosted some... guests who preferred discretion... that's all! I really don't know more!",
            "您...您说的对...Webb先生死了...": "You... you're right... Mr. Webb is dead...",
            "我不是来找你麻烦的。我只想知道真相。": "I'm not here to make trouble for you. I just want the truth.",
            "我...我确实看到过一些人。但我不知道他们是谁...都是晚上来，从后门进。Webb先生会亲自接待。": "I... I did see some people. But I don't know who they were... came at night, through the back door. Mr. Webb would personally receive them.",
            "昨晚呢？": "What about last night?",
            "昨晚...我好像看到有个穿风衣的人...10点多从后门进去的。但我没看清脸...真的...": "Last night... I think I saw someone in a trench coat... went in through the back door around 10. But I couldn't see the face... really...",
            "Webb死了，规矩也变了。": "Webb is dead, the rules have changed.",
            "我...我真的不知道更多了！Webb先生对我不错，给的工钱也公道，我...我不想卷进这些事...求您了...": "I... I really don't know more! Mr. Webb treated me well, paid fair wages, I... I don't want to get involved... please...",
            "如果你想起什么，来找我。": "If you remember anything, come find me.",
            "好的，好的，Brennan先生。如果我想起什么，一定...一定告诉您。我会...会注意的。有任何线索，我马上联系您。": "Okay, okay, Mr. Brennan. If I remember anything, I'll... definitely tell you. I'll... be careful. Any leads, I'll contact you right away.",

            # ===== 证词对话翻译 =====
            "Rosa，我们又见面了。": "Rosa, we meet again.",
            "Brennan先生……我正在打扫，马上就走，您如果——": "Mr. Brennan... I'm cleaning, about to leave, if you—",
            "谢谢你告诉我真相。接下来照常工作，照顾 Miguel。Morrison由我来处理。": "Thank you for telling me the truth. Keep working as usual, take care of Miguel. I'll handle Morrison.",
            "他会杀我的……": "He'll kill me...",
            "他不会知道是你说的。Rosa，我对付过比他更脏的人。": "He won't know you talked. Rosa, I've dealt with dirtier people than him.",
            "谢谢您……我真的别无选择……": "Thank you... I really had no choice...",
            "你为 Miguel 已经做得够多。剩下的交给我。": "You've done enough for Miguel. Leave the rest to me.",

            # ===== NPC info 翻译 =====
            # Rosa
            "蓝月亮歌舞厅清洁工，夜班23:00-01:00": "Blue Moon Club cleaner, night shift 23:00-01:00",
            "单身母亲，儿子Miguel患病需要昂贵药物": "Single mother, son Miguel is ill and needs expensive medication",
            "声称在地下室酒窖工作（实际在后台走廊）": "Claims to work in basement cellar (actually in backstage corridor)",
            "表现紧张，似乎在隐瞒什么": "Appears nervous, seems to be hiding something",
            "被Morrison威胁配合栽赃，内心充满恐惧和愧疚": "Threatened by Morrison to cooperate in framing, filled with fear and guilt",
            "在Zack的劝说下说出真相": "Told the truth after Zack's persuasion",
            "看到Vivian从歌舞厅往会客室方向走，携带可疑物品": "Saw Vivian walking from the club towards the parlor, carrying suspicious items",
            # Morrison
            "芝加哥警局警探，负责Webb谋杀案调查": "Chicago Police detective, in charge of Webb murder investigation",
            "声称00:30接到电话，01:00到达现场": "Claims received call at 00:30, arrived at scene at 01:00",
            "行为可疑，似乎早就知道现场情况": "Suspicious behavior, seems to know the scene beforehand",
            "欠疤面Tony 5000美元赌债，被威胁家人安全": "Owes Scarface Tony $5,000 gambling debt, family threatened",
            "被神秘人\"Whale\"收买，负责栽赃Zack": "Bribed by mysterious 'Whale' to frame Zack",
            "从家到酒吧只需15分钟，但声称用了30分钟": "Takes only 15 minutes from home to bar, but claims it took 30",
            "案发前一天领用了便携式现场勘验箱": "Checked out portable crime scene kit the day before",
            # Tommy
            "蓝月亮歌舞厅经理，住在酒吧楼上": "Blue Moon Club manager, lives above the bar",
            "23:30听到一声枪响，与平时黑帮火拼不同": "Heard a gunshot at 23:30, different from usual gang fights",
            "知道Rosa的排班表和员工动向": "Knows Rosa's schedule and staff movements",
            "对Webb的其他生意闪烁其词": "Evasive about Webb's other business",
            "确认从Morrison家到酒吧开车最多15分钟": "Confirms driving from Morrison's home to bar takes 15 minutes max",
            "深度参与Webb的勒索网络，负责伪装账目": "Deeply involved in Webb's extortion network, handles falsified accounts",
            "知道Webb和Vivian的关系，以及Webb与Rita的新关系": "Knows Webb and Vivian's relationship, and Webb's new relationship with Rita",
            # Vivian
            "确认Morrison到达时间是01:00": "Confirms Morrison arrived at 01:00",
            "对Webb的死表面悲伤，内心复杂": "Outwardly sad about Webb's death, inwardly conflicted",
            "听到Webb和某人争吵，对方劝Webb不要接危险生意": "Heard Webb arguing with someone who warned him against dangerous business",
            "被Webb利用，被迫接近有钱客人套取秘密进行勒索": "Exploited by Webb, forced to get close to wealthy patrons to extract secrets for blackmail",
            "发现Webb和Rita交往后内心充满仇恨": "Filled with hatred after discovering Webb was seeing Rita",
            "有杀人动机和计划，但最终没有动手": "Had motive and plan to kill, but ultimately didn't act",
            "11点05分去会客室，门锁着被赶走": "Went to parlor at 11:05, door was locked and was turned away",
            # James
            "酒吧厨师，与Webb关系密切": "Bar cook, close relationship with Webb",
            "曾警告Webb不要接危险生意": "Once warned Webb against taking dangerous business",
            "拥有Webb会客室的钥匙": "Has key to Webb's parlor",
            "从Webb办公室拿走古董水壶并伪装": "Took antique kettle from Webb's office and disguised it",
            "知道\"Whale\"的危险性": "Knows how dangerous 'Whale' is",
            "对Vivian的情况表现出同情": "Shows sympathy for Vivian's situation",
            "为了妻子Anna和未出生孩子的合法身份，被Whale收买": "Bribed by Whale for his wife Anna and unborn child's legal status",
            "11点杀死Webb，11点05分模仿声音骗走Vivian，11点30分制造假枪声": "Killed Webb at 11:00, imitated voice to drive away Vivian at 11:05, created fake gunshot at 11:30",
            "最终自杀身亡": "Eventually committed suicide",
            # Anna
            "James的妻子，怀有身孕": "James's wife, pregnant",
            "不知道丈夫的犯罪行为": "Unaware of husband's crime",
            "允许Zack进入卧室查看书籍": "Allowed Zack to enter bedroom to check books",
            "James自杀后，向Zack提供了Webb办公室保险柜的密码": "After James's suicide, provided Zack with Webb's office safe combination",
            # Mrs. Morrison
            "Morrison的妻子，对丈夫深信不疑": "Morrison's wife, has complete faith in husband",
            "确认Morrison离家时间为00:30": "Confirms Morrison left home at 00:30",
            "允许Zack和Emma进入书房调查": "Allowed Zack and Emma to investigate the study",
        }

        return translations.get(cn_text, cn_text)

    # ==================== 转换方法 ====================

    def convert_npc(self) -> List[dict]:
        """转换NPC表（v2.1版本 - 27个字段）

        字段结构：
        - 基础信息: id, cnName, enName, role, path1~3
        - 证词: TestimonyCount, cnTestimony, enTestimony
        - 描述: cnDescribe, enDescribe
        - 独立info: info1~info6 (格式: 中文/英文)
        - 指证组1: ifExposeInfo1, cnNewInfo1, enNewInfo1
        - 指证组2: ifExposeInfo2, cnNewInfo2, enNewInfo2
        - 关系图: npcPosX, npcPosY, npcRelation, npcRelationParaCn, npcRelationParaEn
        """
        records = []
        npcs_data = self.npcs.get('npcs', {})

        for npc_id, npc in npcs_data.items():
            # 收集描述信息
            cn_desc = npc.get('description', '')
            # 如果有英文描述就用，没有就用中文（源数据应该有英文）
            en_desc = npc.get('description_en', '')
            if not en_desc and cn_desc:
                en_desc = cn_desc  # 保留中文，等待手动翻译

            # 收集info信息 (从各循环合并，只取前6个，格式：中文/英文)
            info_list = []
            info = npc.get('info', {})
            info_en = npc.get('info_en', {})  # 英文info（如果有）
            # 按loop1, loop2, loop3...顺序排序
            for loop_key in sorted(info.keys()):
                loop_info = info[loop_key]
                loop_info_en = info_en.get(loop_key, []) if info_en else []
                if isinstance(loop_info, list):
                    for idx, cn_info in enumerate(loop_info):
                        # 尝试获取对应的英文，没有则使用翻译
                        en_info = loop_info_en[idx] if idx < len(loop_info_en) else self.translate(cn_info)
                        info_list.append(f"{cn_info}/{en_info}")
                # 只需要前6个，超出的不要
                if len(info_list) >= 6:
                    info_list = info_list[:6]
                    break

            record = {
                'id': npc_id,
                'cnName': npc.get('name_cn', ''),
                'enName': npc.get('name', ''),
                'role': npc.get('role', ''),
                'path1': '',  # 待补充
                'path2': '',  # 待补充
                'path3': '',  # 待补充
                'TestimonyCount': '',  # 待补充
                'cnTestimony': '',  # 待补充
                'enTestimony': '',  # 待补充
                'cnDescribe': cn_desc,
                'enDescribe': en_desc,
                # 独立info字段 (格式: 中文/英文，只取前6个)
                'info1': info_list[0] if len(info_list) > 0 else '',
                'info2': info_list[1] if len(info_list) > 1 else '',
                'info3': info_list[2] if len(info_list) > 2 else '',
                'info4': info_list[3] if len(info_list) > 3 else '',
                'info5': info_list[4] if len(info_list) > 4 else '',
                'info6': info_list[5] if len(info_list) > 5 else '',
                # 指证组1
                'ifExposeInfo1': '',  # 待补充 (1-6)
                'cnNewInfo1': '',  # 待补充 (格式: 原文/新文)
                'enNewInfo1': '',  # 待补充
                # 指证组2
                'ifExposeInfo2': '',  # 待补充 (1-6)
                'cnNewInfo2': '',  # 待补充
                'enNewInfo2': '',  # 待补充
                # 关系图字段 (从源数据读取)
                'npcPosX': npc.get('npcPosX', ''),
                'npcPosY': npc.get('npcPosY', ''),
                'npcRelation': npc.get('npcRelation', ''),
                'npcRelationParaCn': npc.get('npcRelationParaCn', ''),
                'npcRelationParaEn': npc.get('npcRelationParaEn', '')
            }
            records.append(record)

        return records

    def convert_scene(self) -> List[dict]:
        """转换场景表（完整10个字段）"""
        records = []
        scenes_data = self.scenes.get('scenes', {})

        for scene_id, scene in scenes_data.items():
            record = {
                'sceneId': scene_id,
                'sceneName': scene.get('name', ''),
                'sceneNameEn': scene.get('name_en', ''),
                'sceneType': 'dialogue',  # 默认类型
                'backgroundImage': f"Art/Scenes/{scene.get('asset_id', '')}.png",
                'backgroundMusic': '',  # 待补充
                'ambientSound': '',  # 待补充
                'unlockCondition': '',  # 待补充
                'npcsPresent': '',  # 待补充
                '备注': scene.get('description', ''),
            }
            records.append(record)

        return records

    def convert_item(self) -> List[dict]:
        """转换物品表（v2.0版本 - 21个字段）

        字段结构：
        - 基础: id, cnName, enName, itemType
        - 交互: canCollected, canAnalyzed, analysedEvidence, canCombined, combineParameter0/1
        - 描述: cnDescribe1(详细), cnDescribe2(简单/列表), cnDescribe3(分析后)
        - 描述: enDescribe1, enDescribe2, enDescribe3
        - 资源: path1, path2, path3
        - 事件: script(固定JudgeCondition), parameter(事件ID)
        """
        records = []
        evidences_data = self.evidences.get('evidences', {})

        for ev_id, ev in evidences_data.items():
            desc = ev.get('description', {})
            # cnDescribe1 = 详细描述 (initial)
            initial_desc = desc.get('initial', '') if isinstance(desc, dict) else str(desc)
            # cnDescribe2 = 简单描述/列表用 (brief)
            brief_desc = desc.get('brief', '') if isinstance(desc, dict) else ''

            analysis = ev.get('analysis', {})
            # cnDescribe3 = 分析后描述 (result_description)
            analysis_desc = analysis.get('result_description', '') if isinstance(analysis, dict) else ''

            # 判断是否需要分析
            has_analysis = 'analysis' in ev and isinstance(ev['analysis'], dict)

            # 英文描述：优先使用源数据中的英文，没有则使用翻译
            en_desc1 = ev.get('description_en', '')
            if not en_desc1 and initial_desc:
                en_desc1 = self.translate(initial_desc)  # 使用翻译

            en_desc2 = ''
            if brief_desc:
                en_desc2 = self.translate(brief_desc)  # 使用翻译

            en_desc3 = ''
            if analysis_desc:
                # 尝试从analysis获取英文结果描述
                en_analysis_desc = analysis.get('result_description_en', '') if isinstance(analysis, dict) else ''
                en_desc3 = en_analysis_desc if en_analysis_desc else self.translate(analysis_desc)  # 使用翻译

            record = {
                'id': ev_id,
                'cnName': ev.get('name', ''),
                'enName': ev.get('name_en', '') or ev.get('name', ''),  # 英文名优先使用源数据
                'itemType': ev.get('type', 'item'),
                'canCollected': 1 if ev.get('type') in ['item', 'clue', 'note'] else 0,
                'canAnalyzed': 1 if has_analysis else 0,
                'analysedEvidence': '',  # 待补充 (分析后变成的物品ID)
                'canCombined': 0,  # 待补充
                'combineParameter0': '',  # 待补充
                'combineParameter1': '',  # 待补充
                # 描述字段 (v2.0语义)
                'cnDescribe1': initial_desc,  # 详细描述
                'cnDescribe2': brief_desc,  # 简单描述/列表用
                'cnDescribe3': analysis_desc if has_analysis else '',  # 分析后描述
                'enDescribe1': en_desc1,  # 详细描述(英文)
                'enDescribe2': en_desc2,  # 简单描述(英文)
                'enDescribe3': en_desc3,  # 分析后描述(英文)
                # 资源路径 (从evidences.yaml的asset_id读取)
                'path1': ev.get('asset_id', ''),  # 从asset_id读取
                'path2': '',  # 待补充
                'path3': '',  # 待补充
                # 事件触发
                'script': 'JudgeCondition' if ev.get('parameter') else '',  # 固定填写
                'parameter': ev.get('parameter', ''),  # 事件ID
            }
            records.append(record)

        return records

    def convert_talk(self) -> List[dict]:
        """转换对话表（完整19个字段）- 支持自动ID生成和branches分支"""
        records = []
        npcs_data = self.npcs.get('npcs', {})

        # 重置段落计数器
        self.npc_segment_counter = {}

        # ID 分配日志
        id_allocation_log = []

        # ===== 第一遍：收集每个 section 的第一句 ID =====
        section_first_ids = {}  # key: "dialog_name/section_key" -> first_id

        for loop_num, loop_dialogs in sorted(self.dialogs.items()):
            sorted_dialogs = self._sort_dialog_files(loop_dialogs)

            for dialog_name, dialog_data in sorted_dialogs:
                if not dialog_data:
                    continue

                main_npc = self._get_dialog_npc(dialog_data, dialog_name)
                npc_number = self._get_npc_number(main_npc)

                if npc_number == 0:
                    continue

                for section_key, section in dialog_data.items():
                    if not isinstance(section, dict) or 'lines' not in section:
                        continue

                    segment = self._get_next_segment(loop_num, main_npc)

                    # 找到第一句非 player_choice 的对话
                    for line in section.get('lines', []):
                        if line.get('speaker') != 'player_choice':
                            first_id = int(f"{npc_number}{str(segment).zfill(3)}001")
                            section_first_ids[f"{dialog_name}/{section_key}"] = first_id
                            break

        # 重置段落计数器（第二遍重新计算）
        self.npc_segment_counter = {}

        # ===== 第二遍：生成记录，处理 branches =====
        for loop_num, loop_dialogs in sorted(self.dialogs.items()):
            sorted_dialogs = self._sort_dialog_files(loop_dialogs)

            for dialog_name, dialog_data in sorted_dialogs:
                if not dialog_data:
                    continue

                main_npc = self._get_dialog_npc(dialog_data, dialog_name)
                npc_number = self._get_npc_number(main_npc)

                if npc_number == 0:
                    print(f"  [WARN] 无法确定NPC: loop{loop_num}/{dialog_name}.yaml")
                    continue

                for section_key, section in dialog_data.items():
                    if not isinstance(section, dict) or 'lines' not in section:
                        continue

                    segment = self._get_next_segment(loop_num, main_npc)
                    sentence_num = 1

                    id_allocation_log.append({
                        'loop': loop_num,
                        'file': dialog_name,
                        'section': section_key,
                        'npc': main_npc,
                        'segment': segment,
                        'id_prefix': f"{npc_number}{str(segment).zfill(3)}"
                    })

                    lines = section.get('lines', [])
                    for i, line in enumerate(lines):
                        speaker_id = line.get('speaker', '')

                        # 处理 player_choice：生成 branches
                        if speaker_id == 'player_choice':
                            options = line.get('options', [])
                            if not options or len(records) == 0:
                                continue

                            # 获取前一句记录，添加 branches 属性
                            prev_record = records[-1]

                            # 收集选项信息
                            option_texts = []
                            option_reply_ids = []
                            target_section_ids = []

                            for opt_idx, opt in enumerate(options[:3]):  # 最多3个选项
                                option_texts.append(opt.get('text', ''))
                                next_section = opt.get('next_section', '')
                                target_key = f"{dialog_name}/{next_section}"
                                target_id = section_first_ids.get(target_key, 0)
                                target_section_ids.append(str(target_id))

                                # 生成主角复述选项的对话 ID
                                reply_id = int(f"{npc_number}{str(segment).zfill(3)}{str(sentence_num).zfill(3)}")
                                option_reply_ids.append(reply_id)
                                sentence_num += 1

                            # 设置前一句的 branches 参数
                            prev_record['script'] = 'branches'
                            prev_record['next'] = '/'.join(target_section_ids)

                            if len(option_texts) > 0:
                                prev_record['ParameterStr0'] = option_texts[0]
                                prev_record['ParameterInt0'] = option_reply_ids[0]
                            if len(option_texts) > 1:
                                prev_record['ParameterStr1'] = option_texts[1]
                                prev_record['ParameterInt1'] = option_reply_ids[1]
                            if len(option_texts) > 2:
                                prev_record['ParameterStr2'] = option_texts[2]
                                prev_record['ParameterInt2'] = option_reply_ids[2]

                            # 生成主角复述选项的对话行
                            zack_info = npcs_data.get('NPC101', {})
                            for opt_idx, opt in enumerate(options[:3]):
                                reply_id = option_reply_ids[opt_idx]
                                target_key = f"{dialog_name}/{opt.get('next_section', '')}"
                                target_id = section_first_ids.get(target_key, 0)

                                opt_cn_text = opt.get('text', '')
                                opt_en_text = opt.get('text_en', '')
                                # 前600条记录尝试翻译
                                if not opt_en_text and len(records) < 600:
                                    opt_en_text = self.translate(opt_cn_text)
                                elif not opt_en_text:
                                    opt_en_text = opt_cn_text

                                reply_record = {
                                    'id': reply_id,
                                    'step': sentence_num - len(options) + opt_idx,
                                    'speakType': 2,
                                    'waitTime': 0,
                                    'IdSpeaker': 'NPC101',
                                    'cnSpeaker': zack_info.get('name_cn', '查克'),
                                    'enSpeaker': zack_info.get('name', 'Zack'),
                                    'cnWords': opt_cn_text,
                                    'enWords': opt_en_text,
                                    'next': str(target_id),
                                    'script': '',
                                    'ParameterStr0': '',
                                    'ParameterStr1': '',
                                    'ParameterStr2': '',
                                    'ParameterInt0': 0,
                                    'ParameterInt1': 0,
                                    'ParameterInt2': 0,
                                    'imagePath': '',
                                    'voicePath': '',
                                }
                                records.append(reply_record)

                            continue

                        # 普通对话处理
                        speaker_info = npcs_data.get(speaker_id, {})

                        if speaker_id == 'narration':
                            speak_type = 1
                        elif speaker_id:
                            speak_type = 2
                        else:
                            speak_type = 3

                        if line.get('talk_id'):
                            talk_id = line['talk_id']
                        else:
                            talk_id = int(f"{npc_number}{str(segment).zfill(3)}{str(sentence_num).zfill(3)}")

                        # 计算下一句ID（如果是段落最后一句，next=0且script=end）
                        is_last_line = (i == len(lines) - 1) or all(
                            l.get('speaker') == 'player_choice' for l in lines[i+1:]
                        )

                        if is_last_line:
                            next_id = '0'
                            script_value = 'end'
                        else:
                            # 找到下一个非player_choice的行
                            next_sentence_num = sentence_num + 1
                            for j in range(i + 1, len(lines)):
                                if lines[j].get('speaker') != 'player_choice':
                                    next_id = str(int(f"{npc_number}{str(segment).zfill(3)}{str(next_sentence_num).zfill(3)}"))
                                    break
                            else:
                                next_id = '0'
                            script_value = line.get('script', '')

                        cn_words = line.get('text', '')
                        en_words = line.get('text_en', '')
                        # 前600条记录尝试翻译
                        if not en_words and len(records) < 600:
                            en_words = self.translate(cn_words)
                        elif not en_words:
                            en_words = cn_words

                        record = {
                            'id': talk_id,
                            'step': sentence_num,
                            'speakType': speak_type,
                            'waitTime': line.get('wait_time', 0),
                            'IdSpeaker': speaker_id if speaker_id != 'narration' else '',
                            'cnSpeaker': speaker_info.get('name_cn', '旁白' if speaker_id == 'narration' else ''),
                            'enSpeaker': speaker_info.get('name', 'Narration' if speaker_id == 'narration' else ''),
                            'cnWords': cn_words,
                            'enWords': en_words,
                            'next': next_id,
                            'script': script_value if not is_last_line else 'end',
                            'ParameterStr0': line.get('ParameterStr0', ''),
                            'ParameterStr1': line.get('ParameterStr1', ''),
                            'ParameterStr2': line.get('ParameterStr2', ''),
                            'ParameterInt0': line.get('ParameterInt0', 0),
                            'ParameterInt1': line.get('ParameterInt1', 0),
                            'ParameterInt2': line.get('ParameterInt2', 0),
                            'imagePath': '',
                            'voicePath': '',
                        }
                        records.append(record)
                        sentence_num += 1

        # 打印 ID 分配日志
        if id_allocation_log:
            print("\n  📋 Talk ID 分配 (每个 section 一个段落):")
            for log in id_allocation_log:
                print(f"     loop{log['loop']}/{log['file']}.yaml/{log['section']} → {log['npc']} 段落{log['segment']} (ID前缀: {log['id_prefix']})")

        return records

    def convert_testimony(self) -> List[dict]:
        """转换证词表（完整9个字段）- 暂只输出第一循环"""
        records = []
        npcs_data = self.npcs.get('npcs', {})

        # 暂时只输出第一循环
        loop_num = '1'
        loop_dialogs = self.dialogs.get(loop_num, {})

        # 主要从 accusation.yaml 提取证词
        accusation = loop_dialogs.get('accusation', {})
        if not accusation:
            return records

        base_id = int(f"3{loop_num.zfill(2)}1001")
        step = 1
        evidence_counter = 1  # 证词序号计数器

        for section_key, section in accusation.items():
            if not isinstance(section, dict) or 'lines' not in section:
                continue

            for line in section['lines']:
                speaker_id = line.get('speaker', '')
                speaker_info = npcs_data.get(speaker_id, {})

                cn_text = line.get('text', '')
                # 英文：优先使用text_en，没有则使用翻译
                en_text = line.get('text_en', '')
                if not en_text and cn_text:
                    en_text = self.translate(cn_text)  # 使用翻译

                # 判断是否是证词（包含关键证据信息的对话）
                # 检测关键词：时间、地点、看到、听到、声音、人物等
                is_evidence = any(kw in cn_text for kw in [
                    '点', '时', '分', '看到', '听到', '声音', '枪', '人', '进', '出',
                    '门', '房间', '现场', '证明', '目击', '当时', '那天', '晚上'
                ])

                # 生成提取内容（简化版证词摘要）
                cn_extracted = ''
                en_extracted = ''
                if is_evidence and len(cn_text) > 10:
                    # 提取关键信息（取前30个字符作为摘要）
                    cn_extracted = cn_text[:30] + '...' if len(cn_text) > 30 else cn_text
                    en_extracted = self.translate(cn_extracted)  # 使用翻译

                record = {
                    'id': base_id + step - 1,
                    'speakerName': speaker_info.get('name_cn', ''),
                    'speakerNameEn': speaker_info.get('name', ''),
                    'cnWords': cn_text,
                    'enWords': en_text,
                    'ifIgnore': 0,  # 默认显示
                    'ifEvidence': evidence_counter if is_evidence else 0,
                    'cnExracted': cn_extracted,
                    'enExracted': en_extracted,
                }
                records.append(record)
                step += 1
                if is_evidence:
                    evidence_counter += 1

        return records

    def convert_task_chapter_loop(self) -> List[dict]:
        """转换任务表 - 子页签1: Task(章节)(循环)

        字段结构（7个字段）：
        - ID: 章节循环ID (格式: Task + 章节 + 循环, 如 Task11)
        - ChapterTextCn: 主章节任务文本（中文），如"谁杀了Webb？"
        - ChapterTextEn: 主章节任务文本（英文）
        - PhaseGoalCn: 当前循环主任务（中文）
        - PhaseGoalEn: 当前循环主任务（英文）
        - Condition: 完成条件（证据ID，多个用/分隔）
        - TaskID: 关联的具体任务ID（多个用/分隔）
        """
        records = []
        chapter = 1  # 当前是第1章

        # 第一章主线：谁杀了Webb？
        chapter_text_cn = "谁杀了Webb？"
        chapter_text_en = "Who killed Webb?"

        # 循环目标的英文翻译映射
        phase_goal_translations = {
            "到底是谁把我迷晕了，还想把杀人的罪名扣在我头上？": "Who knocked me out and tried to frame me for murder?",
            "Morrison为什么要陷害我？背后是谁指使？": "Why is Morrison framing me? Who's behind this?",
            "Webb的真正生意是什么？": "What was Webb's real business?",
            "James和Webb的真实关系是什么？": "What's the real relationship between James and Webb?",
            "Vivian和Webb的关系是什么？她为什么要杀Webb？": "What's Vivian's relationship with Webb? Why would she kill him?",
            "James如何杀害Webb并伪造时间线？": "How did James kill Webb and fake the timeline?",
        }

        for loop_num in sorted(self.loops.keys()):
            loop_data = self.loops[loop_num]

            # 循环目标（从 investigation_target 读取）
            phase_goal_cn = loop_data.get('investigation_target', '')
            phase_goal_en = loop_data.get('investigation_target_en', '')
            if not phase_goal_en and phase_goal_cn:
                phase_goal_en = phase_goal_translations.get(phase_goal_cn, phase_goal_cn)

            # 完成条件（从 evidence_required 读取，如果有）
            condition = ''
            evidence_required = loop_data.get('evidence_required', [])
            if evidence_required:
                condition = '/'.join(evidence_required)

            # 关联任务ID（暂时留空，需要从 TaskDetail 对应）
            task_id = ''

            record = {
                'ID': f"Task{chapter}{loop_num}",
                'ChapterTextCn': chapter_text_cn,
                'ChapterTextEn': chapter_text_en,
                'PhaseGoalCn': phase_goal_cn,
                'PhaseGoalEn': phase_goal_en,
                'Condition': condition,
                'TaskID': task_id,
            }
            records.append(record)

        return records

    def convert_task_detail(self) -> List[dict]:
        """转换任务表 - 子页签2: Task(详细)

        字段结构（5个字段）：
        - ID: 任务唯一ID (格式: Task + 3位序号, 如 Task001)
        - TaskType: 任务类型 (CurrentGoal / Doubt / SideCase)
        - Condition: 触发条件（Event ID）
        - TaskContentCn: 任务内容（中文）
        - TaskContentEn: 任务内容（英文）
        """
        records = []
        task_counter = 1

        # 英文翻译映射
        translations = {
            # 调查目标
            "到底是谁把我迷晕了，还想把杀人的罪名扣在我头上？": "Who knocked me out and tried to frame me for murder?",
            "Morrison为什么要陷害我？背后是谁指使？": "Why is Morrison framing me? Who's behind this?",
            "Webb的真正生意是什么？": "What was Webb's real business?",
            "James和Webb的真实关系是什么？": "What's the real relationship between James and Webb?",
            "Vivian和Webb的关系是什么？她为什么要杀Webb？": "What's Vivian's relationship with Webb? Why would she kill him?",
            "James如何杀害Webb并伪造时间线？": "How did James kill Webb and fake the timeline?",
            # 核心谎言
            "Rosa声称一直在地下室酒窖工作，什么都没看到": "Rosa claims she was working in the basement cellar and saw nothing",
            "Morrison声称凌晨0点半接警后立即赶到现场": "Morrison claims he rushed to the scene immediately after receiving the call at 12:30 AM",
            "Tommy声称Webb只做私酒生意，没有其他业务": "Tommy claims Webb only dealt in bootleg liquor, nothing else",
            "James声称自己只是普通厨师，和Webb只是普通雇佣关系": "James claims he's just an ordinary cook with a simple employer relationship with Webb",
            "Vivian声称自己和Webb只是普通的雇佣关系": "Vivian claims she and Webb had just a normal employer relationship",
            "James声称案发时一直在厨房，有不在场证明": "James claims he was in the kitchen the whole time with an alibi",
            # 下一目标
            "Morrison为何要陷害我？": "Why is Morrison framing me?",
            "Webb在勒索谁？为什么Whale要杀他？": "Who was Webb blackmailing? Why did Whale want him dead?",
            "James和Webb的关系是什么？Webb勒索的\"Whale\"是谁？": "What's James's relationship with Webb? Who is the 'Whale' Webb was blackmailing?",
            "Vivian和Webb的关系是什么？她知道Whale是谁吗？": "What's Vivian's relationship with Webb? Does she know who Whale is?",
            "James是真凶吗？他如何伪造时间线？": "Is James the real killer? How did he fake the timeline?",
            "Whale是谁？他为什么要杀Webb？": "Who is Whale? Why did he want Webb dead?",
        }

        for loop_num in sorted(self.loops.keys()):
            loop_data = self.loops[loop_num]

            # 从循环数据提取任务内容
            # 1. CurrentGoal: 当前行动目标（从 investigation_target 或 objectives 提取）
            investigation_target = loop_data.get('investigation_target', '')
            if investigation_target:
                record = {
                    'ID': f"Task{str(task_counter).zfill(3)}",
                    'TaskType': 'CurrentGoal',
                    'Condition': '',
                    'TaskContentCn': investigation_target,
                    'TaskContentEn': loop_data.get('investigation_target_en', translations.get(investigation_target, investigation_target)),
                }
                records.append(record)
                task_counter += 1

            # 2. Doubt: 疑点（从 core_lie 提取）
            core_lie = loop_data.get('core_lie', '')
            if core_lie:
                record = {
                    'ID': f"Task{str(task_counter).zfill(3)}",
                    'TaskType': 'Doubt',
                    'Condition': '',
                    'TaskContentCn': core_lie,
                    'TaskContentEn': loop_data.get('core_lie_en', translations.get(core_lie, core_lie)),
                }
                records.append(record)
                task_counter += 1

            # 3. 下一目标作为支线或额外疑点
            ending = loop_data.get('ending', {})
            next_objective = ending.get('next_objective', '')
            if next_objective:
                record = {
                    'ID': f"Task{str(task_counter).zfill(3)}",
                    'TaskType': 'Doubt',
                    'Condition': '',
                    'TaskContentCn': next_objective,
                    'TaskContentEn': ending.get('next_objective_en', translations.get(next_objective, next_objective)),
                }
                records.append(record)
                task_counter += 1

        return records

    def convert_timeline(self) -> List[dict]:
        """转换时间线表（8个字段）

        字段结构：
        - id: 时间线ID (格式: TL + 章节 + 循环 + 序号)
        - Chapter: 章节
        - Loop: 循环
        - time: 时间点 (如 "23:30")
        - cnEvent: 中文事件描述
        - enEvent: 英文事件描述
        - npcId: 关联NPC ID
        - sceneId: 关联场景ID
        """
        records = []
        chapter = 1  # 当前是第1章

        # 基于案情时间线生成（第一章的时间线）
        # 这是根据故事情节预设的时间线
        timeline_events = [
            {'loop': 1, 'time': '23:00', 'cn': 'Rosa开始夜班清洁工作', 'en': 'Rosa starts night shift cleaning', 'npc': 'NPC103', 'scene': 'SC1001'},
            {'loop': 1, 'time': '23:30', 'cn': 'Tommy听到一声枪响', 'en': 'Tommy hears a gunshot', 'npc': 'NPC105', 'scene': 'SC1003'},
            {'loop': 1, 'time': '00:30', 'cn': 'Morrison声称接到电话', 'en': 'Morrison claims to receive a phone call', 'npc': 'NPC104', 'scene': ''},
            {'loop': 1, 'time': '01:00', 'cn': 'Morrison到达现场', 'en': 'Morrison arrives at the scene', 'npc': 'NPC104', 'scene': 'SC1004'},
            {'loop': 2, 'time': '00:30', 'cn': 'Morrison离开家前往酒吧', 'en': 'Morrison leaves home for the bar', 'npc': 'NPC104', 'scene': ''},
            {'loop': 3, 'time': '23:00', 'cn': '音乐声极响，无法听到其他声音', 'en': 'Music is extremely loud, impossible to hear other sounds', 'npc': '', 'scene': 'SC1010'},
            {'loop': 4, 'time': '23:00', 'cn': 'Rosa看到James用钥匙进入Webb会客室', 'en': 'Rosa sees James entering Webb parlor with a key', 'npc': 'NPC107', 'scene': 'SC1004'},
            {'loop': 5, 'time': '23:00', 'cn': 'Vivian从歌舞厅往会客室方向走', 'en': 'Vivian walks from the club towards the parlor', 'npc': 'NPC106', 'scene': 'SC1010'},
            {'loop': 6, 'time': '23:00', 'cn': 'James杀死Webb', 'en': 'James kills Webb', 'npc': 'NPC107', 'scene': 'SC1004'},
            {'loop': 6, 'time': '23:05', 'cn': 'James模仿Webb声音骗走Vivian', 'en': 'James imitates Webb\'s voice to drive away Vivian', 'npc': 'NPC107', 'scene': 'SC1004'},
            {'loop': 6, 'time': '23:30', 'cn': 'James制造假枪声', 'en': 'James creates fake gunshot sound', 'npc': 'NPC107', 'scene': 'SC1004'},
        ]

        for idx, event in enumerate(timeline_events, 1):
            record = {
                'id': f"TL{chapter}{event['loop']}{str(idx).zfill(2)}",
                'Chapter': chapter,
                'Loop': event['loop'],
                'time': event['time'],
                'cnEvent': event['cn'],
                'enEvent': event['en'],
                'npcId': event['npc'],
                'sceneId': event['scene'],
            }
            records.append(record)

        return records

    def convert_chapter_config(self) -> List[dict]:
        """转换章节配置表（v1.1版本 - 5个字段）

        字段结构：
        - id: 游戏场景状态ID (格式: CC + 章节 + 循环, 如 CC11)
        - Chapter: 所属章节 (1-9)
        - Loop: 所属循环 (1-6)
        - SceneID: 可访问场景ID (多个用/分隔)
        - TaskID: 关联Task表宏观ID (如 Task11)
        """
        records = []

        # 从 loops 数据生成章节配置
        # 假设当前是第1章 (Unit1)
        chapter = 1

        for loop_num in sorted(self.loops.keys()):
            loop_data = self.loops[loop_num]

            # 收集该循环的可访问场景ID (从 scenes_overview 获取)
            scene_ids = []
            scenes_overview = loop_data.get('scenes_overview', [])
            for scene_info in scenes_overview:
                # 只收集状态为 accessible 的场景
                if scene_info.get('status') == 'accessible':
                    scene_id = scene_info.get('scene', '')
                    if scene_id:
                        scene_ids.append(scene_id)

            record = {
                'id': f"CC{chapter}{loop_num}",
                'Chapter': chapter,
                'Loop': int(loop_num),
                'SceneID': '/'.join(scene_ids) if scene_ids else '',
                'TaskID': f"Task{chapter}{loop_num}",
            }
            records.append(record)

        return records

    # ==================== 输出方法 ====================

    def save_yaml(self, data: List[dict], filename: str, meta: dict):
        """保存为yaml格式"""
        output = {
            '_meta': meta,
            'data': data
        }

        path = STORY_OUTPUT / f"{filename}.yaml"
        with open(path, 'w', encoding='utf-8') as f:
            yaml.dump(output, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

        print(f"  📄 {path.name}: {len(data)} 条记录")
        return path

    def save_excel(self, data: List[dict], filename: str, meta: dict):
        """保存为Excel格式（带Luban表头）"""
        if not data:
            print(f"  [WARN] {filename}: 无数据，跳过")
            return None

        # 使用 openpyxl 直接创建带表头的 Excel
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # 第1行：##var + 字段名
        ws.append(['##var'] + meta['var'])
        # 第2行：##type + 类型定义
        ws.append(['##type'] + meta['type'])
        # 第3行：## + 字段描述
        ws.append(['##'] + meta['desc'])

        # 第4行起：数据（第一列留空，与表头对齐）
        for record in data:
            row = [''] + [record.get(field, '') for field in meta['var']]
            ws.append(row)

        # 保存
        path = STORY_OUTPUT / f"{filename}.xlsx"
        try:
            wb.save(path)
            print(f"  [OK] {path.name}: {len(data)} 条记录")
        except PermissionError:
            print(f"  ⚠️ {path.name}: 文件被占用，跳过Excel保存")
            return None
        return path

    def save_talk_excel_colored(self, data: List[dict], filename: str, meta: dict):
        """保存Talk表为带颜色的Excel格式"""
        if not data:
            print(f"  [WARN] {filename}: 无数据，跳过")
            return None

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        # 定义颜色
        LIGHT_BLUE = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        LIGHT_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        DARK_GRAY = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        wb = Workbook()
        ws = wb.active

        # 第1行：##var + 字段名
        ws.append(['##var'] + meta['var'])
        # 第2行：##type + 类型定义
        ws.append(['##type'] + meta['type'])
        # 第3行：## + 字段描述
        ws.append(['##'] + meta['desc'])

        # 解析ID获取NPC编号和段落号
        def parse_talk_id(talk_id):
            """从 talk_id 解析 NPC编号和段落号"""
            id_str = str(talk_id)
            if len(id_str) == 7:
                # 格式: NNXXYYY (如 2001001)
                npc_num = int(id_str[0])
                segment = int(id_str[1:4])
            elif len(id_str) == 8:
                # 格式: NNXXXYYY (如 10001001)
                npc_num = int(id_str[0:2])
                segment = int(id_str[2:5])
            else:
                npc_num = 0
                segment = 0
            return npc_num, segment

        # 第4行起：数据（带颜色）
        current_row = 4
        prev_npc_num = None
        col_count = len(meta['var']) + 1  # +1 for first empty column

        for record in data:
            talk_id = record.get('id', 0)
            npc_num, segment = parse_talk_id(talk_id)

            # 检查是否需要插入 NPC 分隔行
            if prev_npc_num is not None and npc_num != prev_npc_num:
                # 插入空行作为分隔
                ws.append([''] * col_count)
                # 给分隔行上色（深灰）
                for col in range(1, col_count + 1):
                    ws.cell(row=current_row, column=col).fill = DARK_GRAY
                current_row += 1

            # 写入数据行
            row = [''] + [record.get(field, '') for field in meta['var']]
            ws.append(row)

            # 根据段落号奇偶决定颜色
            fill_color = LIGHT_BLUE if segment % 2 == 1 else LIGHT_YELLOW
            for col in range(1, col_count + 1):
                ws.cell(row=current_row, column=col).fill = fill_color

            prev_npc_num = npc_num
            current_row += 1

        # 保存
        path = STORY_OUTPUT / f"{filename}.xlsx"
        try:
            wb.save(path)
            print(f"  [OK] {path.name}: {len(data)} 条记录 (带颜色)")
        except PermissionError:
            print(f"  ⚠️ {path.name}: 文件被占用，跳过Excel保存")
            return None
        return path

    def copy_to_unity(self):
        """复制Excel到Unity目录"""
        print("\n📦 复制到Unity目录...")

        if not UNITY_OUTPUT.exists():
            print(f"  ⚠️ Unity目录不存在: {UNITY_OUTPUT}")
            return

        for xlsx in STORY_OUTPUT.glob("*.xlsx"):
            target = UNITY_OUTPUT / xlsx.name
            try:
                shutil.copy2(xlsx, target)
                print(f"  ✅ {xlsx.name} -> {target}")
            except PermissionError:
                print(f"  ⚠️ 权限拒绝，跳过: {xlsx.name} (文件可能被打开)")
            except Exception as e:
                print(f"  ⚠️ 复制失败: {xlsx.name} - {e}")

    # ==================== 主流程 ====================

    def run(self, tables: List[str] = None):
        """执行转换"""
        print("=" * 50)
        print("🚀 配表小助手 - 开始转换")
        print("=" * 50)

        # 加载数据
        self.load_all_data()

        # 定义表配置（严格按照规则文档的完整字段）
        table_configs = {
            'NPCStaticData': {
                'converter': self.convert_npc,
                'meta': {
                    'var': ['id', 'cnName', 'enName', 'role', 'path1', 'path2', 'path3',
                            'TestimonyCount', 'cnTestimony', 'enTestimony',
                            'cnDescribe', 'enDescribe',
                            'info1', 'info2', 'info3', 'info4', 'info5', 'info6',
                            'ifExposeInfo1', 'cnNewInfo1', 'enNewInfo1',
                            'ifExposeInfo2', 'cnNewInfo2', 'enNewInfo2',
                            'npcPosX', 'npcPosY', 'npcRelation', 'npcRelationParaCn', 'npcRelationParaEn'],
                    'type': ['string', 'string', 'string', 'string', 'string', 'string', 'string',
                             'int', 'string', 'string',
                             'string', 'string',
                             'string', 'string', 'string', 'string', 'string', 'string',
                             'int', 'string', 'string',
                             'int', 'string', 'string',
                             'float', 'float', 'string', 'string', 'string'],
                    'desc': ['NPC ID', '中文名', '英文名', '角色类型', '资源路径1', '资源路径2', '资源路径3',
                             '证词数量', '中文证词', '英文证词',
                             '中文描述', '英文描述',
                             '人物信息1', '人物信息2', '人物信息3', '人物信息4', '人物信息5', '人物信息6',
                             '指证info编号1', '指证后中文1', '指证后英文1',
                             '指证info编号2', '指证后中文2', '指证后英文2',
                             '关系图X坐标', '关系图Y坐标', '关联NPC', '关系描述(中)', '关系描述(英)'],
                }
            },
            'SceneConfig': {
                'converter': self.convert_scene,
                'meta': {
                    'var': ['sceneId', 'sceneName', 'sceneNameEn', 'sceneType',
                            'backgroundImage', 'backgroundMusic', 'ambientSound',
                            'unlockCondition', 'npcsPresent', '备注'],
                    'type': ['string', 'string', 'string', 'string',
                             'string', 'string', 'string',
                             'string', 'string', 'string'],
                    'desc': ['场景ID', '中文场景名', '英文场景名', '场景类型',
                             '背景图路径', '背景音乐', '环境音效',
                             '解锁条件', '场景NPC', '备注'],
                }
            },
            'ItemStaticData': {
                'converter': self.convert_item,
                'meta': {
                    'var': ['id', 'cnName', 'enName', 'itemType',
                            'canCollected', 'canAnalyzed', 'analysedEvidence', 'canCombined',
                            'combineParameter0', 'combineParameter1',
                            'cnDescribe1', 'cnDescribe2', 'cnDescribe3',
                            'enDescribe1', 'enDescribe2', 'enDescribe3',
                            'path1', 'path2', 'path3', 'script', 'parameter'],
                    'type': ['string', 'string', 'string', 'string',
                             'int', 'int', 'string', 'int',
                             'string', 'string',
                             'string', 'string', 'string',
                             'string', 'string', 'string',
                             'string', 'string', 'string', 'string', 'string'],
                    'desc': ['物品ID', '中文名', '英文名', '物品类型',
                             '可收集(1是0否)', '可分析(1是0否)', '分析后物品ID', '可合并(1是0否)',
                             '合并参数0', '合并参数1',
                             '中文详细描述', '中文简单描述', '中文分析后描述',
                             '英文详细描述', '英文简单描述', '英文分析后描述',
                             '资源路径1', '资源路径2', '资源路径3', '脚本方法', '事件ID'],
                }
            },
            'Talk': {
                'converter': self.convert_talk,
                'meta': {
                    'var': ['id', 'step', 'speakType', 'waitTime', 'IdSpeaker', 'cnSpeaker', 'enSpeaker',
                            'cnWords', 'enWords', 'next', 'script',
                            'ParameterStr0', 'ParameterStr1', 'ParameterStr2',
                            'ParameterInt0', 'ParameterInt1', 'ParameterInt2',
                            'imagePath', 'voicePath'],
                    'type': ['int', 'int', 'int', 'float', 'string', 'string', 'string',
                             'string', 'string', 'string', 'string',
                             'string', 'string', 'string',
                             'int', 'int', 'int',
                             'string', 'string'],
                    'desc': ['对话ID', '步骤', '对话类型', '等待时间', '说话人ID', '中文名', '英文名',
                             '中文台词', '英文台词', '下一句ID', '脚本类型',
                             '字符串参数0', '字符串参数1', '字符串参数2',
                             '整数参数0', '整数参数1', '整数参数2',
                             '头像路径', '语音路径'],
                }
            },
            'Testimony': {
                'converter': self.convert_testimony,
                'meta': {
                    'var': ['id', 'speakerName', 'speakerNameEn', 'cnWords', 'enWords',
                            'ifIgnore', 'ifEvidence', 'cnExracted', 'enExracted'],
                    'type': ['int', 'string', 'string', 'string', 'string',
                             'int', 'int', 'string', 'string'],
                    'desc': ['证词ID', '说话人中文名', '说话人英文名', '中文内容', '英文内容',
                             '是否隐藏', '证词序号', '中文提取', '英文提取'],
                }
            },
            'ChapterConfig': {
                'converter': self.convert_chapter_config,
                'meta': {
                    'var': ['id', 'Chapter', 'Loop', 'SceneID', 'TaskID'],
                    'type': ['string', 'int', 'int', 'string', 'string'],
                    'desc': ['游戏状态ID', '所属章节', '所属循环', '可访问场景ID', '任务ID'],
                }
            },
            'TaskChapterLoop': {
                'converter': self.convert_task_chapter_loop,
                'meta': {
                    'var': ['ID', 'ChapterTextCn', 'ChapterTextEn', 'PhaseGoalCn', 'PhaseGoalEn', 'Condition', 'TaskID'],
                    'type': ['string', 'string', 'string', 'string', 'string', 'string', 'string'],
                    'desc': ['章节循环ID', '主章节任务(中)', '主章节任务(英)', '循环目标(中)', '循环目标(英)', '完成条件', '关联任务ID'],
                }
            },
            'TaskDetail': {
                'converter': self.convert_task_detail,
                'meta': {
                    'var': ['ID', 'TaskType', 'Condition', 'TaskContentCn', 'TaskContentEn'],
                    'type': ['string', 'string', 'string', 'string', 'string'],
                    'desc': ['任务ID', '任务类型', '触发条件', '任务内容(中)', '任务内容(英)'],
                }
            },
            'Timeline': {
                'converter': self.convert_timeline,
                'meta': {
                    'var': ['id', 'Chapter', 'Loop', 'time', 'cnEvent', 'enEvent', 'npcId', 'sceneId'],
                    'type': ['string', 'int', 'int', 'string', 'string', 'string', 'string', 'string'],
                    'desc': ['时间线ID', '章节', '循环', '时间点', '中文事件', '英文事件', '关联NPC', '关联场景'],
                }
            },
        }

        # 表名别名映射（支持简写）
        table_aliases = {
            'npc': 'NPCStaticData',
            'scene': 'SceneConfig',
            'item': 'ItemStaticData',
            'talk': 'Talk',
            'testimony': 'Testimony',
            'chapter': 'ChapterConfig',
            'task': 'TaskChapterLoop',  # 默认指向子页签1
            'taskloop': 'TaskChapterLoop',
            'taskdetail': 'TaskDetail',
            'timeline': 'Timeline',
        }

        # 确定要处理的表
        if tables:
            # 将别名转换为完整表名
            resolved_tables = []
            for t in tables:
                if t in table_aliases:
                    resolved_tables.append(table_aliases[t])
                elif t in table_configs:
                    resolved_tables.append(t)
            table_configs = {k: v for k, v in table_configs.items() if k in resolved_tables}

        # 执行转换
        print("\n📝 生成配置表...")
        results = {}

        for name, config in table_configs.items():
            data = config['converter']()
            self.save_yaml(data, name, config['meta'])
            # Talk表使用带颜色的保存方法
            if name == 'Talk':
                self.save_talk_excel_colored(data, name, config['meta'])
            else:
                self.save_excel(data, name, config['meta'])
            results[name] = len(data)

        # 复制到Unity
        self.copy_to_unity()

        # 输出统计
        print("\n" + "=" * 50)
        print("✅ 转换完成！")
        print("=" * 50)
        print("\n📊 统计:")
        for name, count in results.items():
            print(f"  {name}: {count} 条")

        return results


if __name__ == '__main__':
    import sys

    converter = ConfigTableConverter()

    # 支持命令行参数指定表
    if len(sys.argv) > 1:
        tables = sys.argv[1:]
        converter.run(tables)
    else:
        converter.run()
